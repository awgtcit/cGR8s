"""
Report Generation Service – PDF via WeasyPrint, Excel via openpyxl.
"""
import logging
import os
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)


class ReportGenerator:
    """Generates PDF and Excel reports."""

    def __init__(self, output_dir: str = 'reports', template_dir: str = 'app/reports/templates'):
        self.output_dir = output_dir
        self.template_dir = template_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def generate_pdf(self, template_name: str, context: dict, filename: str = None) -> str:
        """Render an HTML template and convert to PDF via WeasyPrint."""
        from flask import render_template_string
        from weasyprint import HTML

        template_path = os.path.join(self.template_dir, template_name)
        with open(template_path, 'r', encoding='utf-8') as f:
            template_content = f.read()

        html_content = render_template_string(template_content, **context)

        if not filename:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filename = f'{template_name.replace(".html", "")}_{ts}.pdf'

        filepath = os.path.join(self.output_dir, filename)
        HTML(string=html_content).write_pdf(filepath)
        logger.info('PDF report generated: %s', filepath)
        return filepath

    def generate_excel(self, data: list, headers: list, sheet_name: str = 'Report',
                       filename: str = None) -> str:
        """Generate an Excel report from tabular data."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0D6B3C', end_color='0D6B3C', fill_type='solid')

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit column widths
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        if not filename:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            filename = f'{sheet_name}_{ts}.xlsx'

        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)
        logger.info('Excel report generated: %s', filepath)
        return filepath
