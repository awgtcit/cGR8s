"""QA Workflow module – QA analysis entry and update."""
from datetime import datetime, timedelta
from io import BytesIO
from flask import Blueprint, render_template, request, redirect, url_for, g, jsonify, send_file
from app.auth.decorators import require_auth, require_permission
from app.config.constants import Permissions, AuditAction, ProcessOrderStatus
from app.repositories import (
    ProcessOrderRepository, QAAnalysisRepository, QAUpdateRepository,
    FGCodeRepository, TargetWeightResultRepository, NPLResultRepository,
    NPLInputRepository, MachineRepository,
)
from app.models.qa import QAAnalysis
from app.utils.helpers import paginate_args, flash_success, flash_error
from app.audit import AuditLogger

bp = Blueprint('qa', __name__, template_folder='templates')

# Dropdown options for QA entry (select_field expects list of {value, label} dicts)
COMPANY_OPTIONS = [
    {'value': 'ALW', 'label': 'ALW'},
    {'value': 'UTC', 'label': 'UTC'},
    {'value': 'NGF', 'label': 'NGF'},
]
def _get_mc_options(db):
    """Load active machines from the DB for the MC dropdown."""
    machines = MachineRepository(db).get_all()
    return [
        {'value': m.machine_code, 'label': m.machine_code}
        for m in machines if m.is_active
    ]


@bp.route('/')
@require_auth
@require_permission(Permissions.QA_VIEW)
def index():
    page, per_page = paginate_args(request.args)
    repo = QAAnalysisRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page)
    # Build PO number lookup for display (single bulk query)
    from app.repositories import ProcessOrderRepository
    po_repo = ProcessOrderRepository(g.db)
    po_ids = list({qa.process_order_id for qa in result.get('items', []) if qa.process_order_id})
    po_lookup = po_repo.get_po_number_map(po_ids)
    return render_template('qa/index.html', po_lookup=po_lookup, **result)


@bp.route('/pending')
@require_auth
@require_permission(Permissions.QA_VIEW)
def pending():
    """Process orders pending QA."""
    po_repo = ProcessOrderRepository(g.db)
    page, per_page = paginate_args(request.args)
    result = po_repo.get_paginated(
        page=page, per_page=per_page,
        filters={'status': ProcessOrderStatus.QA_PENDING.value},
    )
    return render_template('qa/pending.html', **result)


def _build_prefill(po, fg, po_repo):
    """Pre-populate QA entry form with target values and computed dates."""
    data = {}

    # NPL Date = Prod Date + 1 day
    if po.process_date:
        data['npl_date'] = (po.process_date + timedelta(days=1)).strftime('%Y-%m-%d')

    # QR Date = last run's process_date for same FG (excluding current PO)
    if po.fg_code_id:
        previous_pos = po_repo.get_by_fg_code(po.fg_code_id)
        last_run_po = next((p for p in previous_pos if p.id != po.id), None)
        if last_run_po and last_run_po.process_date:
            data['qr_date'] = last_run_po.process_date.strftime('%Y-%m-%d')

    # Pre-fill with target values from FG Code master
    if fg:
        targets = {
            'total_cig_length': fg.cig_length,
            'circumference_mean': fg.circumference_mean,
            'cig_pdo': fg.cig_pdo,
            'tip_vf': fg.tip_ventilation,
            'w_ntm': fg.ntm_wt_mean,
            'filter_pd_mean': fg.filter_pd,
            'cig_hardness': fg.cig_hardness,
            'cig_corr_hardness': fg.cig_corrected_hardness,
            'maker_moisture': fg.maker_moisture,
            'pack_ov': fg.pack_ov,
            'ssi': fg.ssi,
            'lamina_cpi': fg.lamina_cpi,
            'filling_power': fg.filling_power,
            'plug_wrap_cu': fg.plug_wrap_cu,
            'plug_length': fg.plug_length,
            'tow': fg.tow_used,
        }
        for key, val in targets.items():
            if val is not None:
                data[key] = val

    return data


@bp.route('/enter/<process_order_id>', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.QA_ENTER)
def enter(process_order_id):
    po_repo = ProcessOrderRepository(g.db)
    po = po_repo.get_by_id(process_order_id)
    if not po:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Process Order', process_order_id)

    # Get FG code for reference info display
    fg = None
    if po.fg_code_id:
        fg = FGCodeRepository(g.db).get_by_id(po.fg_code_id)

    if request.method == 'POST':
        data = request.form.to_dict()
        data['process_order_id'] = process_order_id
        data['status'] = 'pending_review'

        # Convert numeric fields
        numeric_fields = [
            'qa_w_cig', 'qa_w_tob', 'qa_moisture', 'qa_nicotine', 'qa_tar', 'qa_co',
            'pack_ov', 'lamina_cpi', 'filling_power', 'filling_power_corr',
            'maker_moisture', 'ssi', 'pan_pct', 'total_cig_length',
            'circumference_mean', 'circumference_sd', 'cig_dia',
            'tip_vf', 'tip_vf_sd', 'filter_pd_mean', 'w_ntm',
            'plug_wrap_cu', 'cig_pdo', 'cig_hardness', 'cig_corr_hardness',
            'loose_shorts', 'plug_length',
        ]
        for field in numeric_fields:
            val = data.get(field, '').strip() if data.get(field) else ''
            data[field] = float(val) if val else None

        # Keep string fields as-is: tow, company, mc, notes, qr_date, npl_date
        for str_field in ['tow', 'company', 'mc', 'notes']:
            if str_field in data and not data[str_field].strip():
                data[str_field] = None

        # Date fields
        for date_field in ['qr_date', 'npl_date']:
            if date_field in data and not data[date_field].strip():
                data[date_field] = None

        qa_repo = QAAnalysisRepository(g.db)
        from app.models.qa import QAAnalysis
        qa = qa_repo.create(QAAnalysis(**data))

        # Move PO to QA_UPDATED
        po_repo.update(po.id, {'status': ProcessOrderStatus.QA_UPDATED.value})
        g.db.commit()

        AuditLogger.log(AuditAction.CREATE, 'QAAnalysis',
                        entity_id=qa.id, after_value=data, module='qa')
        flash_success('QA data entered')
        return redirect(url_for('qa.detail', id=qa.id))

    return render_template('qa/enter.html', po=po, fg=fg, data=_build_prefill(po, fg, po_repo),
                           mc_options=_get_mc_options(g.db), company_options=COMPANY_OPTIONS)


@bp.route('/<id>')
@require_auth
@require_permission(Permissions.QA_VIEW)
def detail(id):
    repo = QAAnalysisRepository(g.db)
    qa = repo.get_by_id(id)
    if not qa:
        from app.utils.errors import NotFoundError
        raise NotFoundError('QA Analysis', id)
    return render_template('qa/detail.html', qa=qa)


@bp.route('/<id>/approve', methods=['POST'])
@require_auth
@require_permission(Permissions.QA_APPROVE)
def approve(id):
    qa_repo = QAAnalysisRepository(g.db)
    qa = qa_repo.get_by_id(id)
    if not qa:
        from app.utils.errors import NotFoundError
        raise NotFoundError('QA Analysis', id)

    qa_repo.update(id, {'status': 'approved'})

    # Create QA update record
    update_repo = QAUpdateRepository(g.db)
    update_repo.create({
        'qa_analysis_id': id,
        'process_order_id': qa.process_order_id,
        'action': 'approved',
        'notes': request.form.get('notes', ''),
    })
    g.db.commit()

    AuditLogger.log(AuditAction.APPROVE, 'QAAnalysis', entity_id=id, module='qa')
    flash_success('QA analysis approved')
    return redirect(url_for('qa.detail', id=id))


@bp.route('/<id>/reject', methods=['POST'])
@require_auth
@require_permission(Permissions.QA_APPROVE)
def reject(id):
    qa_repo = QAAnalysisRepository(g.db)
    qa = qa_repo.get_by_id(id)
    if not qa:
        from app.utils.errors import NotFoundError
        raise NotFoundError('QA Analysis', id)

    qa_repo.update(id, {'status': 'rejected'})

    update_repo = QAUpdateRepository(g.db)
    update_repo.create({
        'qa_analysis_id': id,
        'process_order_id': qa.process_order_id,
        'action': 'rejected',
        'notes': request.form.get('notes', ''),
    })
    g.db.commit()

    AuditLogger.log(AuditAction.REJECT, 'QAAnalysis', entity_id=id, module='qa')
    flash_success('QA analysis rejected')
    return redirect(url_for('qa.detail', id=id))


@bp.route('/data-grid')
@require_auth
@require_permission(Permissions.QA_VIEW)
def data_grid():
    """Excel-like data grid with Production Data, QA Analysis, Daily Operation tabs."""
    po_repo = ProcessOrderRepository(g.db)
    fg_repo = FGCodeRepository(g.db)
    tw_repo = TargetWeightResultRepository(g.db)
    npl_repo = NPLResultRepository(g.db)
    npl_input_repo = NPLInputRepository(g.db)
    qa_repo = QAAnalysisRepository(g.db)

    show_all = request.args.get('all', '0') == '1'

    # Fetch ALL orders for QR date lookup (prev production date for same FG)
    all_orders = po_repo.get_all()
    all_orders.sort(key=lambda o: o.process_date or datetime.min, reverse=True)

    # Build previous production date map: {fg_code_id: [dates sorted desc]}
    fg_dates = {}
    for o in all_orders:
        if o.fg_code_id and o.process_date:
            fg_dates.setdefault(o.fg_code_id, []).append((o.id, o.process_date))

    orders = all_orders if show_all else all_orders[:10]

    # Build lookups — bulk fetch to avoid N+1 queries
    po_ids = [o.id for o in orders]
    fg_ids = list({o.fg_code_id for o in orders if o.fg_code_id})
    fg_map = {}
    for fid in fg_ids:
        fg_obj = fg_repo.get_by_id(fid)
        if fg_obj:
            fg_map[fid] = fg_obj

    tw_map = tw_repo.get_by_process_orders(po_ids)
    npl_map = npl_repo.get_by_process_orders(po_ids)
    npl_input_map = npl_input_repo.get_by_process_orders(po_ids)
    qa_map = qa_repo.get_by_process_orders(po_ids)

    rows = []
    for po in orders:
        fg = fg_map.get(po.fg_code_id)

        # Compute QR Date = latest production date BEFORE current date for same FG (SKU+Blend)
        qr_date = None
        if po.fg_code_id and po.fg_code_id in fg_dates and po.process_date:
            for pid, pdate in fg_dates[po.fg_code_id]:
                if pid != po.id and pdate < po.process_date:
                    qr_date = pdate
                    break  # first one before current date (sorted desc = most recent)

        # Compute NPL Date = Prod Date + 1 day
        npl_date = (po.process_date + timedelta(days=1)) if po.process_date else None

        row = {
            'po': po,
            'fg': fg,
            'tw': tw_map.get(po.id),
            'npl': npl_map.get(po.id),
            'npl_input': npl_input_map.get(po.id),
            'qa': qa_map.get(po.id),
            'prev_qa': None,
            'qr_date': qr_date,
            'npl_date': npl_date,
            # UID = date + PO number + blend
            'uid': f"{po.process_date.strftime('%Y%m%d') if po.process_date else ''}-{po.process_order_number}-{fg.blend_code or '' if fg else ''}",
        }
        rows.append(row)

    # ── Auto-fill: for rows without QA, get latest QA from previous production with same FG ──
    no_qa_fg_ids = list({r['po'].fg_code_id for r in rows
                         if not r['qa'] and r['po'].fg_code_id})
    if no_qa_fg_ids:
        no_qa_po_ids = [r['po'].id for r in rows if not r['qa']]
        prev_qa_map = qa_repo.get_latest_by_fg_code_ids(no_qa_fg_ids, exclude_po_ids=no_qa_po_ids)
        for r in rows:
            if not r['qa'] and r['po'].fg_code_id and r['po'].fg_code_id in prev_qa_map:
                r['prev_qa'] = prev_qa_map[r['po'].fg_code_id]

    # MC and Company options for the paste modal dropdowns
    mc_options = _get_mc_options(g.db)
    return render_template('qa/data_grid.html', rows=rows, show_all=show_all,
                           mc_options=mc_options, company_options=COMPANY_OPTIONS)


@bp.route('/data-grid/export-excel')
@require_auth
@require_permission(Permissions.QA_VIEW)
def export_excel():
    """Download a combined Excel workbook matching the cGR8s.xlsm format."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # ── Reuse data_grid query logic ──
    po_repo = ProcessOrderRepository(g.db)
    fg_repo = FGCodeRepository(g.db)
    tw_repo = TargetWeightResultRepository(g.db)
    npl_repo = NPLResultRepository(g.db)
    npl_input_repo = NPLInputRepository(g.db)
    qa_repo = QAAnalysisRepository(g.db)

    all_orders = po_repo.get_all()
    all_orders.sort(key=lambda o: o.process_date or datetime.min, reverse=True)

    fg_dates = {}
    for o in all_orders:
        if o.fg_code_id and o.process_date:
            fg_dates.setdefault(o.fg_code_id, []).append((o.id, o.process_date))

    orders = all_orders
    po_ids = [o.id for o in orders]
    fg_ids = list({o.fg_code_id for o in orders if o.fg_code_id})
    fg_map = {}
    for fid in fg_ids:
        fg_obj = fg_repo.get_by_id(fid)
        if fg_obj:
            fg_map[fid] = fg_obj

    tw_map = tw_repo.get_by_process_orders(po_ids)
    npl_map = npl_repo.get_by_process_orders(po_ids)
    npl_input_map = npl_input_repo.get_by_process_orders(po_ids)
    qa_map = qa_repo.get_by_process_orders(po_ids)

    rows = []
    for po in orders:
        fg = fg_map.get(po.fg_code_id)
        qr_date = None
        if po.fg_code_id and po.fg_code_id in fg_dates and po.process_date:
            for pid, pdate in fg_dates[po.fg_code_id]:
                if pid != po.id and pdate < po.process_date:
                    qr_date = pdate
                    break
        npl_date = (po.process_date + timedelta(days=1)) if po.process_date else None
        uid = f"{po.process_date.strftime('%Y%m%d') if po.process_date else ''}-{po.process_order_number}-{fg.blend_code or '' if fg else ''}"
        rows.append({
            'po': po, 'fg': fg,
            'tw': tw_map.get(po.id), 'npl': npl_map.get(po.id),
            'npl_input': npl_input_map.get(po.id), 'qa': qa_map.get(po.id),
            'qr_date': qr_date, 'npl_date': npl_date, 'uid': uid,
        })

    # ── Build Excel workbook ──
    wb = openpyxl.Workbook()
    header_font = Font(bold=True, size=9)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    data_font = Font(size=9)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def _v(obj, attr, fmt=None):
        """Safely get attribute value, optionally formatted."""
        if obj is None:
            return ''
        val = getattr(obj, attr, None)
        if val is None:
            return ''
        if fmt:
            return round(float(val), fmt) if isinstance(fmt, int) else val
        return val

    def _write_sheet(ws, headers, row_builder):
        """Write headers and data rows to a worksheet."""
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        for ri, r in enumerate(rows, 2):
            vals = row_builder(r, ri - 1)
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = data_font
                cell.border = thin_border
        # Auto-width (approximate)
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(len(str(h)) + 2, 10)
        ws.freeze_panes = 'A2'

    # ── Sheet 1: Production Data (109 columns: A-DE) ──
    ws_prod = wb.active
    ws_prod.title = 'Production Data'
    prod_headers = [
        'RID', 'QR Date (D-)', 'Prod. Date (D)', 'NPL Date (D+1)', 'Process Order',
        'SKU', 'SKU Desc', 'SKU GTN', 'Blend Code', 'Blend Desc', 'Blend GTIN', 'Cig. Code', 'UID',
        'NBLD', 'PCU', 'VF', 'FPD', 'MIP',
        'α', 'β', 'γ', 'δ',
        'NC', 'NT', 'NTM', 'TW',
        'TISS', 'TUN', 'RW1', 'RW2', 'RW3', 'RW4', 'RW5', 'WM', 'WP', 'WQ', 'NW', 'NMC', 'NCG', 'T USD', 'M DSP', 'M DST',
        'S1', 'S2', 'TD', 'F', 'NS1', 'NS2', 'NTD', 'NF', 'NTDRY',
        'W DRY', 'W TOB', 'W CIG',
        'TAC', 'TTC', 'NPL%', 'NPL KG',
        'FG+Blend',
        'Pack OV', 'Status', 'Next Row#', 'Revised Weight',
        'N_BLD', 'P_CU', 'V_F', 'F_PD', 'M_IP',
        'S1', 'S2', 'TD', 'F', 'NS1', 'NS2', 'NTD', 'NF', 'NTDRY',
        'W DRY', 'W TOB', 'W CIG', 'Stage No', 'NPL',
    ]

    def prod_row(r, idx):
        po, fg, tw, npl, ni, qa = r['po'], r['fg'], r['tw'], r['npl'], r['npl_input'], r['qa']
        fg_blend = f"{fg.fg_code}, {fg.blend_code}" if fg and fg.fg_code and fg.blend_code else (fg.fg_code if fg else '')
        return [
            idx,  # RID
            r['qr_date'], po.process_date, r['npl_date'], po.process_order_number,
            fg.fg_code if fg else '', fg.brand if fg else '', fg.fg_gtin if fg else '',
            fg.blend_code if fg else '', fg.blend if fg else '', fg.blend_gtin if fg else '',
            fg.cig_code if fg else '', r['uid'],
            # Key Variables
            _v(tw, 'input_n_bld', 4), _v(tw, 'input_p_cu', 4), _v(tw, 'input_t_vnt', 4),
            _v(tw, 'input_f_pd', 4), _v(tw, 'input_m_ip', 4),
            # Constants
            _v(tw, 'input_alpha', 4), _v(tw, 'input_beta', 4),
            _v(tw, 'input_gamma', 4), _v(tw, 'input_delta', 4),
            # NC/NT/NTM/TW
            _v(fg, 'c_plg', 2), _v(fg, 'target_nic', 4),
            _v(fg, 'ntm_wt_mean', 4), _v(tw, 'tw', 2),
            # NPL Input (16)
            _v(ni, 't_iss', 2), _v(ni, 't_un', 2),
            _v(ni, 'l_dst', 2), _v(ni, 'l_win', 2), _v(ni, 'l_flr', 2),
            _v(ni, 'l_srt', 2), _v(ni, 'l_dt', 2),
            _v(ni, 'r_mkg', 2), _v(ni, 'r_pkg', 2), _v(ni, 'r_ndt', 2),
            _v(ni, 'n_w', 2), _v(ni, 'n_mc', 2), _v(ni, 'n_cg', 2),
            _v(ni, 't_usd', 2), _v(ni, 'm_dsp', 2), _v(ni, 'm_dst', 2),
            # TW Results (9)
            _v(tw, 'stage1_dilution', 4), _v(tw, 'stage2_dilution', 4),
            _v(tw, 'total_dilution', 4), _v(tw, 'filtration_pct', 4),
            _v(tw, 'stage1_pacifying_nicotine_demand', 4), _v(tw, 'stage2_pacifying_nicotine_demand', 4),
            _v(tw, 'total_pacifying_nicotine_demand', 4), _v(tw, 'total_filtration_pct', 4),
            _v(tw, 'total_nicotine_demand', 4),
            # TW Output (3)
            _v(tw, 'w_dry', 2), _v(tw, 'w_tob', 2), _v(tw, 'w_cig', 2),
            # NPL Results (4)
            _v(npl, 'tac', 2), _v(npl, 'ttc', 2), _v(npl, 'npl_pct', 2), _v(npl, 'npl_kg', 2),
            # FG+Blend
            fg_blend,
            # Pack OV + Status + Next Row# + Revised Weight
            _v(qa, 'pack_ov', 2), po.status or '', idx, _v(tw, 'tw', 2),
            # Repeated Key Vars (5)
            _v(tw, 'input_n_bld', 4), _v(tw, 'input_p_cu', 4), _v(tw, 'input_t_vnt', 4),
            _v(tw, 'input_f_pd', 4), _v(tw, 'input_m_ip', 4),
            # Repeated TW Results (9)
            _v(tw, 'stage1_dilution', 4), _v(tw, 'stage2_dilution', 4),
            _v(tw, 'total_dilution', 4), _v(tw, 'filtration_pct', 4),
            _v(tw, 'stage1_pacifying_nicotine_demand', 4), _v(tw, 'stage2_pacifying_nicotine_demand', 4),
            _v(tw, 'total_pacifying_nicotine_demand', 4), _v(tw, 'total_filtration_pct', 4),
            _v(tw, 'total_nicotine_demand', 4),
            # Repeated TW Output (3)
            _v(tw, 'w_dry', 2), _v(tw, 'w_tob', 2), _v(tw, 'w_cig', 2),
            # Stage No + NPL
            '',  # Stage No
            _v(npl, 'npl_pct', 2),
        ]

    _write_sheet(ws_prod, prod_headers, prod_row)

    # ── Sheet 2: QA Analysis (38 columns) ──
    ws_qa = wb.create_sheet('QA Analysis')
    qa_headers = [
        'RID', 'QR Date (D-)', 'Prod. Date (D)', 'NPL Date (D+1)', 'Process Order',
        'SKU', 'SKU Desc', 'Blend Code', 'Blend Desc', 'UID',
        'FG+Blend', 'MC', 'Company',
        'Pack OV', 'Lamina CPI', 'Filling Power', 'FP Corr', 'Maker Moist', 'SSI',
        'PAN%', 'Total Cig Length', 'Circ Mean', 'Circ SD', 'Cig Dia',
        'TIP VF', 'TIP VF SD', 'Filter PD', 'W_NTM', 'Plug Wrap CU', 'TOW',
        'Cig PDO', 'Cig Hard', 'Cig Corr Hard', 'Loose Shorts', 'Plug Length',
        'Tob Wt Mean', 'Tob Wt SD', 'Status',
    ]

    def qa_row(r, idx):
        po, fg, qa = r['po'], r['fg'], r['qa']
        fg_blend = f"{fg.fg_code}, {fg.blend_code}" if fg and fg.fg_code and fg.blend_code else (fg.fg_code if fg else '')
        return [
            idx, r['qr_date'], po.process_date, r['npl_date'], po.process_order_number,
            fg.fg_code if fg else '', fg.brand if fg else '',
            fg.blend_code if fg else '', fg.blend if fg else '', r['uid'],
            fg_blend, _v(qa, 'mc'), _v(qa, 'company'),
            _v(qa, 'pack_ov', 2), _v(qa, 'lamina_cpi', 2), _v(qa, 'filling_power', 2),
            _v(qa, 'filling_power_corr', 2), _v(qa, 'maker_moisture', 2), _v(qa, 'ssi', 2),
            _v(qa, 'pan_pct', 2), _v(qa, 'total_cig_length', 2),
            _v(qa, 'circumference_mean', 2), _v(qa, 'circumference_sd', 2), _v(qa, 'cig_dia', 2),
            _v(qa, 'tip_vf', 2), _v(qa, 'tip_vf_sd', 2), _v(qa, 'filter_pd_mean', 2),
            _v(qa, 'w_ntm', 2), _v(qa, 'plug_wrap_cu', 2), _v(qa, 'tow') if qa else '',
            _v(qa, 'cig_pdo', 2), _v(qa, 'cig_hardness', 2), _v(qa, 'cig_corr_hardness', 2),
            _v(qa, 'loose_shorts', 2), _v(qa, 'plug_length', 2),
            '', '',  # Tob Wt Mean/SD
            po.status or '',
        ]

    _write_sheet(ws_qa, qa_headers, qa_row)

    # ── Sheet 3: Daily Operation (60 columns) ──
    ws_daily = wb.create_sheet('Daily Operation')
    daily_headers = [
        'RID', 'QR Date (D-)', 'Prod. Date (D)', 'NPL Date (D+1)', 'Process Order',
        'SKU', 'SKU Desc', 'SKU GTN', 'Blend Code', 'Blend Desc', 'Blend GTIN', 'Cig. Code', 'UID',
        'NBLD', 'PCU', 'VF', 'FPD', 'MIP',
        'α', 'β', 'γ', 'δ',
        'NC', 'NT', 'NTM', 'TW',
        'TP', 'TUN', 'RW1', 'RW2', 'RW3', 'RW4', 'RW5', 'WM', 'WP', 'WQ', 'NW', 'NMC', 'NCG', 'T USD', 'M DSP', 'M DST',
        'S1', 'S2', 'TD', 'F', 'NS1', 'NS2', 'NTD', 'NF', 'NTDRY',
        'W DRY', 'W TOB', 'W CIG',
        'TAC', 'TTC', 'NPL%', 'NPL KG',
        'FG+Blend', 'Status',
    ]

    def daily_row(r, idx):
        po, fg, tw, npl, ni = r['po'], r['fg'], r['tw'], r['npl'], r['npl_input']
        fg_blend = f"{fg.fg_code}, {fg.blend_code}" if fg and fg.fg_code and fg.blend_code else (fg.fg_code if fg else '')
        return [
            idx, r['qr_date'], po.process_date, r['npl_date'], po.process_order_number,
            fg.fg_code if fg else '', fg.brand if fg else '', fg.fg_gtin if fg else '',
            fg.blend_code if fg else '', fg.blend if fg else '', fg.blend_gtin if fg else '',
            fg.cig_code if fg else '', r['uid'],
            _v(tw, 'input_n_bld', 4), _v(tw, 'input_p_cu', 4), _v(tw, 'input_t_vnt', 4),
            _v(tw, 'input_f_pd', 4), _v(tw, 'input_m_ip', 4),
            _v(tw, 'input_alpha', 4), _v(tw, 'input_beta', 4),
            _v(tw, 'input_gamma', 4), _v(tw, 'input_delta', 4),
            _v(fg, 'c_plg', 2), _v(fg, 'target_nic', 4),
            _v(fg, 'ntm_wt_mean', 4), _v(tw, 'tw', 2),
            _v(ni, 't_iss', 2), _v(ni, 't_un', 2),
            _v(ni, 'l_dst', 2), _v(ni, 'l_win', 2), _v(ni, 'l_flr', 2),
            _v(ni, 'l_srt', 2), _v(ni, 'l_dt', 2),
            _v(ni, 'r_mkg', 2), _v(ni, 'r_pkg', 2), _v(ni, 'r_ndt', 2),
            _v(ni, 'n_w', 2), _v(ni, 'n_mc', 2), _v(ni, 'n_cg', 2),
            _v(ni, 't_usd', 2), _v(ni, 'm_dsp', 2), _v(ni, 'm_dst', 2),
            _v(tw, 'stage1_dilution', 4), _v(tw, 'stage2_dilution', 4),
            _v(tw, 'total_dilution', 4), _v(tw, 'filtration_pct', 4),
            _v(tw, 'stage1_pacifying_nicotine_demand', 4), _v(tw, 'stage2_pacifying_nicotine_demand', 4),
            _v(tw, 'total_pacifying_nicotine_demand', 4), _v(tw, 'total_filtration_pct', 4),
            _v(tw, 'total_nicotine_demand', 4),
            _v(tw, 'w_dry', 2), _v(tw, 'w_tob', 2), _v(tw, 'w_cig', 2),
            _v(npl, 'tac', 2), _v(npl, 'ttc', 2), _v(npl, 'npl_pct', 2), _v(npl, 'npl_kg', 2),
            fg_blend, po.status or '',
        ]

    _write_sheet(ws_daily, daily_headers, daily_row)

    # ── Write to buffer and send ──
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"cGR8s_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ── Column mapping: grid column index → (model field, type) ──
_PASTE_COL_MAP = {
    1:  ('qr_date', 'date'),
    3:  ('npl_date', 'date'),
    # 4 = Process Order (used as key, not stored)
    13: ('pack_ov', 'float'),
    14: ('lamina_cpi', 'float'),
    15: ('filling_power', 'float'),
    16: ('filling_power_corr', 'float'),
    17: ('maker_moisture', 'float'),
    18: ('ssi', 'float'),
    19: ('pan_pct', 'float'),
    20: ('total_cig_length', 'float'),
    21: ('circumference_mean', 'float'),
    22: ('circumference_sd', 'float'),
    23: ('cig_dia', 'float'),
    24: ('tip_vf', 'float'),
    25: ('tip_vf_sd', 'float'),
    26: ('filter_pd_mean', 'float'),
    27: ('w_ntm', 'float'),
    28: ('plug_wrap_cu', 'float'),
    29: ('tow', 'string'),
    30: ('cig_pdo', 'float'),
    31: ('cig_hardness', 'float'),
    32: ('cig_corr_hardness', 'float'),
    33: ('loose_shorts', 'float'),
    34: ('plug_length', 'float'),
    35: ('mc', 'string'),
    36: ('company', 'string'),
}

_DATE_FORMATS = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d %b %Y', '%Y%m%d']


def _parse_date(val):
    """Try common date formats; return date object or None."""
    if not val:
        return None
    val = val.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(val):
    """Safely parse a float value from pasted text."""
    if not val:
        return None
    val = val.strip().replace(',', '')
    if val in ('–', '-', ''):
        return None
    try:
        return float(val)
    except ValueError:
        return None


@bp.route('/bulk-paste', methods=['POST'])
@require_auth
@require_permission(Permissions.QA_ENTER)
def bulk_paste():
    """Accept tab-separated rows pasted from Excel, upsert QA records."""
    payload = request.get_json(silent=True)
    if not payload or 'rows' not in payload:
        return jsonify({'ok': False, 'error': 'No data provided'}), 400

    raw_rows = payload['rows']
    if not isinstance(raw_rows, list) or len(raw_rows) == 0:
        return jsonify({'ok': False, 'error': 'Empty data'}), 400

    if len(raw_rows) > 500:
        return jsonify({'ok': False, 'error': 'Too many rows (max 500)'}), 400

    po_repo = ProcessOrderRepository(g.db)
    qa_repo = QAAnalysisRepository(g.db)

    results = {'created': 0, 'updated': 0, 'skipped': [], 'errors': []}

    # ── Step 1: Extract PO numbers and parse all rows ──
    parsed = []
    for row_idx, cols in enumerate(raw_rows, start=1):
        if not isinstance(cols, list) or len(cols) < 5:
            results['errors'].append(f'Row {row_idx}: insufficient columns')
            continue
        po_number = (cols[4] or '').strip()
        if not po_number or po_number in ('\u2013', '-'):
            results['errors'].append(f'Row {row_idx}: missing Process Order')
            continue

        # Parse QA data from column mapping
        qa_data = {}
        for col_idx, (field, ftype) in _PASTE_COL_MAP.items():
            val = cols[col_idx] if col_idx < len(cols) else ''
            if ftype == 'date':
                parsed_val = _parse_date(val)
                if parsed_val is not None:
                    qa_data[field] = parsed_val
            elif ftype == 'float':
                parsed_val = _parse_float(val)
                if parsed_val is not None:
                    qa_data[field] = parsed_val
            elif ftype == 'string':
                clean = (val or '').strip()
                if clean and clean not in ('\u2013', '-'):
                    qa_data[field] = clean

        if not qa_data:
            results['skipped'].append(f'Row {row_idx}: no QA data found')
            continue
        parsed.append((row_idx, po_number, qa_data))

    if not parsed:
        return jsonify({'ok': True, **results})

    # ── Step 2: Bulk pre-fetch POs and existing QA records ──
    unique_po_numbers = list({p[1] for p in parsed})
    po_map = po_repo.get_by_order_numbers(unique_po_numbers)

    matched_po_ids = [po_map[n].id for n in unique_po_numbers if n in po_map]
    qa_map = qa_repo.get_by_process_orders(matched_po_ids)

    # ── Step 3: Upsert with per-row error handling ──
    for row_idx, po_number, qa_data in parsed:
        po = po_map.get(po_number)
        if not po:
            results['skipped'].append(f'Row {row_idx}: PO "{po_number}" not found')
            continue

        try:
            existing_qa = qa_map.get(po.id)
            if existing_qa:
                for key, value in qa_data.items():
                    setattr(existing_qa, key, value)
                existing_qa.updated_at = datetime.now()
                results['updated'] += 1
            else:
                qa_data['process_order_id'] = po.id
                qa_data['status'] = 'pending_review'
                new_qa = qa_repo.create(QAAnalysis(**qa_data))
                qa_map[po.id] = new_qa  # update map for duplicate PO rows
                results['created'] += 1

            if po.status != ProcessOrderStatus.QA_UPDATED.value:
                po.status = ProcessOrderStatus.QA_UPDATED.value
        except Exception as exc:
            results['errors'].append(f'Row {row_idx}: {str(exc)[:80]}')

    g.db.commit()

    AuditLogger.log(AuditAction.CREATE, 'QAAnalysis',
                    after_value={'bulk_paste': True, 'created': results['created'],
                                 'updated': results['updated']}, module='qa')

    return jsonify({'ok': True, **results})
