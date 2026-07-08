"""Master Data module – Blends, Physical Parameters, Calibration Constants, Machines, SKUs, Targets & Limits."""
from flask import Blueprint, render_template, request, redirect, url_for, g, jsonify
from app.auth.decorators import require_auth, require_permission, require_any_permissions
from app.config.constants import Permissions
from app.repositories import (
    BlendMasterRepository, PhysicalParameterRepository,
    CalibrationConstantRepository, FormulaConstantRepository,
    GammaConstantRepository, LookupRepository, FGCodeRepository,
    MachineRepository, SKURepository, TobaccoBlendAnalysisRepository,
)
from app.utils.helpers import paginate_args, flash_success, flash_error
from app.audit import AuditLogger
from app.config.constants import AuditAction

bp = Blueprint('master_data', __name__, template_folder='templates')


@bp.route('/')
@require_auth
@require_permission(Permissions.MASTER_DATA_VIEW)
def index():
    """Master Data root – redirect to blends."""
    return redirect(url_for('master_data.blends'))


# ── Blend Master ──────────────────────────────────────────────────────────

@bp.route('/blends')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_BLENDS)
def blends():
    return render_template('master_data/blends.html')


@bp.route('/blends/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def blend_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = BlendMasterRepository(g.db)
        blend = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'BlendMaster', entity_id=blend.id, after_value=data, module='master_data')
        flash_success('Blend created')
        return redirect(url_for('master_data.blends'))
    return render_template('master_data/blend_form.html', data={}, errors=[])


@bp.route('/blends/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def blend_edit(id):
    repo = BlendMasterRepository(g.db)
    blend = repo.get_by_id(id)
    if not blend:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Blend', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'BlendMaster', entity_id=id, after_value=data, module='master_data')
        flash_success('Blend updated')
        return redirect(url_for('master_data.blends'))
    data = {c.name: getattr(blend, c.name) for c in blend.__table__.columns}
    return render_template('master_data/blend_form.html', data=data, errors=[], blend=blend)


# ── Monthly Blend Nicotine (tobacco_blend_analysis) ─────────────────────

NIC_WET_FACTOR = 0.875  # Nicotine % Wet = Nicotine % Dry (N_BLD) * 0.875


def _nic_editable_months():
    """The two editable periods: (current, previous) as (year, month) tuples."""
    from datetime import date
    today = date.today()
    cur = (today.year, today.month)
    prev = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
    return cur, prev


def _blend_last_used(year, month):
    """{blend_code: last MTC production date in the month}; {} if MTC unreachable."""
    from app.models.fg_code import FGCode
    from app.modules.fg_codes import fetch_mtc_month_brandcode_dates
    try:
        bc_dates = fetch_mtc_month_brandcode_dates(year, month)
    except Exception:
        return {}  # ponytail: MTC offline → no dates, table falls back to code order
    if not bc_dates:
        return {}
    fg_blend = dict(g.db.query(FGCode.fg_code, FGCode.blend_code).filter(
        FGCode.is_deleted == False,  # noqa: E712
        FGCode.blend_code.isnot(None), FGCode.blend_code != ''
    ).all())
    last = {}
    for bc, d in bc_dates.items():
        blend = fg_blend.get(bc)
        if blend and d > last.get(blend, ''):
            last[blend] = d
    return last


@bp.route('/blends/nicotine')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_BLENDS)
def blends_nicotine():
    """All blends with monthly nicotine values; blends produced in the month (MTC) carry last_used."""
    try:
        year, month = int(request.args.get('year', 0)), int(request.args.get('month', 0))
    except ValueError:
        return jsonify({'error': 'Invalid year/month'}), 400
    if not (1 <= month <= 12):
        return jsonify({'error': 'Invalid month'}), 400

    from app.models.blend_master import BlendMaster
    repo = TobaccoBlendAnalysisRepository(g.db)
    saved = repo.get_by_period(year, month)
    last_used = _blend_last_used(year, month)

    from app.models.fg_code import FGCode
    blends = g.db.query(BlendMaster).filter(BlendMaster.is_deleted == False).all()  # noqa: E712
    known = {b.blend_code for b in blends}
    # FG-referenced blend codes missing from Blend Master still get a row (id-less);
    # name/GTIN fall back to what the FG code carries
    orphans = {}
    for c, nm, gt in g.db.query(FGCode.blend_code, FGCode.blend, FGCode.blend_gtin).filter(
        FGCode.is_deleted == False,  # noqa: E712
        FGCode.blend_code.isnot(None), FGCode.blend_code != ''
    ).distinct().all():
        if c not in known:
            cur = orphans.setdefault(c, {'name': None, 'gtin': None})
            cur['name'] = cur['name'] or (nm or '').strip() or None
            cur['gtin'] = cur['gtin'] or gt

    rows = []
    for code, b in [(b.blend_code, b) for b in blends] + [(c, None) for c in orphans]:
        name = (b.blend_name if b else orphans[code]['name']) or None
        # tobacco_blend_analysis.blend_name holds a code or a tobacco name — match both
        keys = [code, (name or '').strip() or None]
        rec = saved.get(keys[0]) or (saved.get(keys[1]) if keys[1] else None)
        prev = None if rec else repo.get_latest_before(keys, year, month)
        src = rec or prev
        rows.append({
            'id': b.id if b else None, 'blend_code': code,
            'blend_name': name,
            'gtin': b.blend_gtin if b else orphans[code]['gtin'],
            'is_active': b.is_active if b else True,
            'last_used': last_used.get(code),
            'nic_dry': src.nic_dry if src else None,
            'nic_wet': src.nic_wet if src else None,
            'dispatch_moisture': src.dispatch_moisture if src else None,
            'saved': bool(rec),
        })
    # Used-this-month first, most recent date first; then the rest by blend code
    rows.sort(key=lambda r: r['blend_code'] or '')
    rows.sort(key=lambda r: r['last_used'] or '', reverse=True)
    return jsonify({'year': year, 'month': month, 'rows': rows})


@bp.route('/blends/nicotine', methods=['POST'])
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_EDIT, Permissions.MASTER_DATA_BLENDS)
def blends_nicotine_save():
    """Upsert monthly nicotine values. Only the current and previous month are editable."""
    data = request.get_json() or {}
    try:
        year, month = int(data.get('year', 0)), int(data.get('month', 0))
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid year/month'}), 400
    if (year, month) not in _nic_editable_months():
        return jsonify({'error': 'Only the current and previous month can be updated'}), 400

    from app.models.blend_master import BlendMaster
    from app.models.fg_code import FGCode
    code_to_name = {b.blend_code: (b.blend_name or '').strip()
                    for b in g.db.query(BlendMaster).filter(BlendMaster.is_deleted == False).all()}  # noqa: E712
    # blend codes missing from Blend Master: use the name the FG code carries
    for c, nm in g.db.query(FGCode.blend_code, FGCode.blend).filter(
        FGCode.is_deleted == False,  # noqa: E712
        FGCode.blend_code.isnot(None), FGCode.blend_code != ''
    ).distinct().all():
        if c and (nm or '').strip() and not code_to_name.get(c):
            code_to_name[c] = nm.strip()

    repo = TobaccoBlendAnalysisRepository(g.db)
    created = updated = 0
    for row in data.get('rows') or []:
        code = str(row.get('blend_code', '')).strip()
        try:
            nic_dry = float(row['nic_dry'])
        except (KeyError, ValueError, TypeError):
            continue  # rows without a valid dry value are skipped
        try:
            moisture = float(row.get('dispatch_moisture')) if row.get('dispatch_moisture') not in (None, '') else None
        except (ValueError, TypeError):
            moisture = None

        values = {'nic_dry': nic_dry, 'nic_wet': nic_dry * NIC_WET_FACTOR,
                  'dispatch_moisture': moisture}
        # Values are per tobacco name — store under the name so all codes sharing it match
        name = code_to_name.get(code) or code
        existing = repo.get_month_value([code, name], year, month)
        if existing:
            repo.update(existing.id, values)
            updated += 1
        else:
            repo.create({'blend_name': name, 'period_year': year, 'period_month': month,
                         'is_active': True, **values})
            created += 1

    g.db.commit()
    AuditLogger.log(AuditAction.UPDATE, 'TobaccoBlendAnalysis', entity_id=None,
                    after_value={'year': year, 'month': month, 'created': created, 'updated': updated},
                    module='master_data')
    return jsonify({'success': True, 'created': created, 'updated': updated})


# ── Physical Parameters ──────────────────────────────────────────────────

@bp.route('/physical-params')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_BLENDS)
def physical_params():
    page, per_page = paginate_args(request.args)
    repo = PhysicalParameterRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page)
    return render_template('master_data/physical_params.html', **result)


@bp.route('/physical-params/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def physical_param_edit(id):
    repo = PhysicalParameterRepository(g.db)
    param = repo.get_by_id(id)
    if not param:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Physical Parameter', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'PhysicalParameter', entity_id=id, after_value=data, module='master_data')
        flash_success('Physical parameter updated')
        return redirect(url_for('master_data.physical_params'))
    data = {c.name: getattr(param, c.name) for c in param.__table__.columns}
    fg_codes = FGCodeRepository(g.db).get_all()
    return render_template('master_data/physical_param_form.html', data=data, param=param, fg_codes=fg_codes, errors={})


# ── Calibration Constants ────────────────────────────────────────────────

@bp.route('/calibration')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_CALIBRATION)
def calibration():
    page, per_page = paginate_args(request.args)
    repo = CalibrationConstantRepository(g.db)
    result = repo.get_paginated_with_fg_search(
        page=page, per_page=per_page,
        search=request.args.get('q', '')
    )
    return render_template('master_data/calibration.html', **result)


@bp.route('/calibration/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def calibration_edit(id):
    repo = CalibrationConstantRepository(g.db)
    cal = repo.get_by_id(id)
    if not cal:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Calibration Constant', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))

        # Sync n_tgt → SKU.nicotine
        if 'n_tgt' in data and cal.fg_code_id:
            try:
                n_tgt_val = float(data['n_tgt']) if data['n_tgt'] else None
            except (ValueError, TypeError):
                n_tgt_val = None
            if n_tgt_val is not None:
                from app.services.nicotine_sync import sync_nicotine
                sync_nicotine(fg_code_id=cal.fg_code_id, n_tgt_val=n_tgt_val)

        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'CalibrationConstant', entity_id=id, after_value=data, module='master_data')
        flash_success('Calibration constant updated')
        return redirect(url_for('master_data.calibration'))
    data = {c.name: getattr(cal, c.name) for c in cal.__table__.columns}
    fg_codes = FGCodeRepository(g.db).get_all()
    return render_template('master_data/calibration_form.html', data=data, constant=cal, fg_codes=fg_codes, errors={})


@bp.route('/calibration/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def calibration_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = CalibrationConstantRepository(g.db)
        cal = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'CalibrationConstant', entity_id=cal.id, after_value=data, module='master_data')
        flash_success('Calibration constant created')
        return redirect(url_for('master_data.calibration'))
    fg_codes = FGCodeRepository(g.db).get_all()
    return render_template('master_data/calibration_form.html', data={}, errors={}, constant=None, fg_codes=fg_codes)


@bp.route('/calibration/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def calibration_delete(id):
    repo = CalibrationConstantRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'CalibrationConstant', entity_id=id, module='master_data')
    flash_success('Calibration constant deleted')
    return redirect(url_for('master_data.calibration'))


@bp.route('/blends/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def blend_delete(id):
    repo = BlendMasterRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'BlendMaster', entity_id=id, module='master_data')
    flash_success('Blend deleted')
    return redirect(url_for('master_data.blends'))


@bp.route('/physical-params/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def physical_param_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = PhysicalParameterRepository(g.db)
        param = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'PhysicalParameter', entity_id=param.id, after_value=data, module='master_data')
        flash_success('Physical parameter created')
        return redirect(url_for('master_data.physical_params'))
    fg_codes = FGCodeRepository(g.db).get_all()
    return render_template('master_data/physical_param_form.html', data={}, errors={}, param=None, fg_codes=fg_codes)


@bp.route('/physical-params/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def physical_param_delete(id):
    repo = PhysicalParameterRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'PhysicalParameter', entity_id=id, module='master_data')
    flash_success('Physical parameter deleted')
    return redirect(url_for('master_data.physical_params'))


# ── Lookups ──────────────────────────────────────────────────────────────

@bp.route('/lookups')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_LOOKUPS)
def lookups():
    repo = LookupRepository(g.db)
    page, per_page = paginate_args(request.args)
    category = request.args.get('category', '')
    result = repo.get_paginated(page=page, per_page=per_page, search=request.args.get('q', ''), search_fields=['code', 'display_name'])
    categories = [r[0] for r in repo.session.query(repo.model_class.category).distinct().all()]
    return render_template('master_data/lookups.html', categories=categories, **result)


@bp.route('/lookups/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def lookup_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = LookupRepository(g.db)
        lk = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'Lookup', entity_id=lk.id, after_value=data, module='master_data')
        flash_success('Lookup created')
        return redirect(url_for('master_data.lookups'))
    return render_template('master_data/lookup_form.html', data={}, errors={}, lookup=None)


@bp.route('/lookups/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def lookup_edit(id):
    repo = LookupRepository(g.db)
    lk = repo.get_by_id(id)
    if not lk:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Lookup', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'Lookup', entity_id=id, after_value=data, module='master_data')
        flash_success('Lookup updated')
        return redirect(url_for('master_data.lookups'))
    data = {c.name: getattr(lk, c.name) for c in lk.__table__.columns}
    return render_template('master_data/lookup_form.html', data=data, lookup=lk, errors={})


@bp.route('/lookups/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def lookup_delete(id):
    repo = LookupRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'Lookup', entity_id=id, module='master_data')
    flash_success('Lookup deleted')
    return redirect(url_for('master_data.lookups'))


# ── Machines ─────────────────────────────────────────────────────────────

@bp.route('/machines')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_MACHINES)
def machines():
    page, per_page = paginate_args(request.args)
    repo = MachineRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['machine_code', 'description'])
    return render_template('master_data/machines.html', **result)


@bp.route('/machines/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def machine_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = MachineRepository(g.db)
        machine = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'Machine', entity_id=machine.id, after_value=data, module='master_data')
        flash_success('Machine created')
        return redirect(url_for('master_data.machines'))
    return render_template('master_data/machine_form.html', data={}, errors={}, machine=None)


@bp.route('/machines/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def machine_edit(id):
    repo = MachineRepository(g.db)
    machine = repo.get_by_id(id)
    if not machine:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Machine', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'Machine', entity_id=id, after_value=data, module='master_data')
        flash_success('Machine updated')
        return redirect(url_for('master_data.machines'))
    data = {c.name: getattr(machine, c.name) for c in machine.__table__.columns}
    return render_template('master_data/machine_form.html', data=data, machine=machine, errors={})


@bp.route('/machines/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def machine_delete(id):
    repo = MachineRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'Machine', entity_id=id, module='master_data')
    flash_success('Machine deleted')
    return redirect(url_for('master_data.machines'))


# ── SKUs ─────────────────────────────────────────────────────────────────

@bp.route('/skus')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_SKUS)
def skus():
    page, per_page = paginate_args(request.args)
    repo = SKURepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['sku_code', 'description'])
    return render_template('master_data/skus.html', **result)


@bp.route('/skus/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def sku_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = SKURepository(g.db)
        sku = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'SKU', entity_id=sku.id, after_value=data, module='master_data')
        flash_success('SKU created')
        return redirect(url_for('master_data.skus'))
    return render_template('master_data/sku_form.html', data={}, errors={}, sku=None)


@bp.route('/skus/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def sku_edit(id):
    repo = SKURepository(g.db)
    sku = repo.get_by_id(id)
    if not sku:
        from app.utils.errors import NotFoundError
        raise NotFoundError('SKU', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))

        # Sync nicotine → CalibrationConstant.n_tgt
        if 'nicotine' in data and sku.sku_code:
            try:
                nic_val = float(data['nicotine']) if data['nicotine'] else None
            except (ValueError, TypeError):
                nic_val = None
            if nic_val is not None:
                from app.services.nicotine_sync import sync_nicotine
                sync_nicotine(fg_code=sku.sku_code, n_tgt_val=nic_val)

        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'SKU', entity_id=id, after_value=data, module='master_data')
        flash_success('SKU updated')
        return redirect(url_for('master_data.skus'))
    data = {c.name: getattr(sku, c.name) for c in sku.__table__.columns}
    return render_template('master_data/sku_form.html', data=data, sku=sku, errors={})


@bp.route('/skus/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def sku_delete(id):
    repo = SKURepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'SKU', entity_id=id, module='master_data')
    flash_success('SKU deleted')
    return redirect(url_for('master_data.skus'))


# ── Tobacco Blend Analysis ───────────────────────────────────────────────

@bp.route('/tobacco-blend-analysis')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_TOBACCO_ANALYSIS)
def tobacco_blend_analysis():
    page, per_page = paginate_args(request.args)
    repo = TobaccoBlendAnalysisRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['blend_name'])
    return render_template('master_data/tobacco_blend_analysis.html', **result)


# ── Formula Constants ────────────────────────────────────────────────────

@bp.route('/formula-constants')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_FORMULA_CONSTANTS)
def formula_constants():
    page, per_page = paginate_args(request.args)
    repo = FormulaConstantRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['name', 'description'])
    return render_template('master_data/formula_constants.html', **result)


@bp.route('/formula-constants/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def formula_constant_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = FormulaConstantRepository(g.db)
        fc = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'FormulaConstant', entity_id=fc.id, after_value=data, module='master_data')
        flash_success('Formula constant created')
        return redirect(url_for('master_data.formula_constants'))
    return render_template('master_data/formula_constant_form.html', data={}, errors={}, constant=None)


@bp.route('/formula-constants/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def formula_constant_edit(id):
    repo = FormulaConstantRepository(g.db)
    fc = repo.get_by_id(id)
    if not fc:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Formula Constant', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'FormulaConstant', entity_id=id, after_value=data, module='master_data')
        flash_success('Formula constant updated')
        return redirect(url_for('master_data.formula_constants'))
    data = {c.name: getattr(fc, c.name) for c in fc.__table__.columns}
    return render_template('master_data/formula_constant_form.html', data=data, constant=fc, errors={})


@bp.route('/formula-constants/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def formula_constant_delete(id):
    repo = FormulaConstantRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'FormulaConstant', entity_id=id, module='master_data')
    flash_success('Formula constant deleted')
    return redirect(url_for('master_data.formula_constants'))


# ── Gamma Constants ──────────────────────────────────────────────────────

@bp.route('/gamma-constants')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_GAMMA_CONSTANTS)
def gamma_constants():
    page, per_page = paginate_args(request.args)
    repo = GammaConstantRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['format', 'selection_criteria'])
    return render_template('master_data/gamma_constants.html', **result)


@bp.route('/gamma-constants/create', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def gamma_constant_create():
    if request.method == 'POST':
        data = request.form.to_dict()
        repo = GammaConstantRepository(g.db)
        gc = repo.create(data)
        g.db.commit()
        AuditLogger.log(AuditAction.CREATE, 'GammaConstant', entity_id=gc.id, after_value=data, module='master_data')
        flash_success('Gamma constant created')
        return redirect(url_for('master_data.gamma_constants'))
    return render_template('master_data/gamma_constant_form.html', data={}, errors={}, constant=None)


@bp.route('/gamma-constants/<id>/edit', methods=['GET', 'POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def gamma_constant_edit(id):
    repo = GammaConstantRepository(g.db)
    gc = repo.get_by_id(id)
    if not gc:
        from app.utils.errors import NotFoundError
        raise NotFoundError('Gamma Constant', id)
    if request.method == 'POST':
        data = request.form.to_dict()
        repo.update(id, data, row_version=int(request.form.get('row_version', 0)))
        g.db.commit()
        AuditLogger.log(AuditAction.UPDATE, 'GammaConstant', entity_id=id, after_value=data, module='master_data')
        flash_success('Gamma constant updated')
        return redirect(url_for('master_data.gamma_constants'))
    data = {c.name: getattr(gc, c.name) for c in gc.__table__.columns}
    return render_template('master_data/gamma_constant_form.html', data=data, constant=gc, errors={})


# ── Reseed Gamma Constants ────────────────────────────────────────────────

@bp.route('/gamma-constants/reseed', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def gamma_constants_reseed():
    """Reseed gamma_constants AND formula_constants with production data."""
    from app.database import get_engine
    from app.services.seed_service import (
        ensure_seed_tables, seed_formula_constants, seed_gamma_constants,
    )

    ensure_seed_tables(get_engine())
    fc_added = seed_formula_constants(g.db)
    added, updated, deactivated = seed_gamma_constants(g.db)
    g.db.commit()

    flash_success(f'Reseeded: Formula constants: {fc_added} added. '
                  f'Gamma: {added} added, {updated} updated, {deactivated} deactivated')
    return redirect(url_for('master_data.gamma_constants'))


@bp.route('/gamma-constants/<id>/delete', methods=['POST'])
@require_auth
@require_permission(Permissions.MASTER_DATA_EDIT)
def gamma_constant_delete(id):
    repo = GammaConstantRepository(g.db)
    repo.soft_delete(id)
    g.db.commit()
    AuditLogger.log(AuditAction.DELETE, 'GammaConstant', entity_id=id, module='master_data')
    flash_success('Gamma constant deleted')
    return redirect(url_for('master_data.gamma_constants'))


# ── Size / CU ────────────────────────────────────────────────────────────

@bp.route('/size-cu')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_SIZE_CU)
def size_cu():
    page, per_page = paginate_args(request.args)
    repo = LookupRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                filters={'category': 'size_cu'},
                                search=request.args.get('q', ''),
                                search_fields=['code', 'display_name'])
    return render_template('master_data/size_cu.html', **result)


# ── KP Tolerance ─────────────────────────────────────────────────────────

@bp.route('/kp-tolerance')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_KP_TOLERANCE)
def kp_tolerance():
    page, per_page = paginate_args(request.args)
    repo = LookupRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                filters={'category': 'kp_tolerance'},
                                search=request.args.get('q', ''),
                                search_fields=['code', 'display_name'])
    return render_template('master_data/kp_tolerance.html', **result)


# ── Plug Length / Cuts ───────────────────────────────────────────────────

@bp.route('/plug-length-cuts')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_PLUG_LENGTH)
def plug_length_cuts():
    page, per_page = paginate_args(request.args)
    repo = LookupRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                filters={'category': 'plug_length_cuts'},
                                search=request.args.get('q', ''),
                                search_fields=['code', 'display_name'])
    return render_template('master_data/plug_length_cuts.html', **result)


# ── App Fields ───────────────────────────────────────────────────────────

@bp.route('/app-fields')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_APP_FIELDS)
def app_fields():
    page, per_page = paginate_args(request.args)
    repo = LookupRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                filters={'category': 'app_fields'},
                                order_by='sort_order',
                                search=request.args.get('q', ''),
                                search_fields=['code', 'display_name'])
    return render_template('master_data/app_fields.html', **result)


# ── Targets & Limits ─────────────────────────────────────────────────────

# Excel header (normalized: ':' stripped, whitespace collapsed, lowercased) → (FGCode field, type)
_TL_FIELD_MAP = {
    'cig code': ('cig_code', str), 'blend code': ('blend_code', str),
    'filter code': ('filter_code', str), 'blend': ('blend', str), 'brand': ('brand', str),
    'format': ('format', str), 'family name': ('family_name', str),
    'fg gtin': ('fg_gtin', str), 'blend gtin': ('blend_gtin', str),
    'circumference mean': ('circumference_mean', float),
    'circumference mean ul': ('circumference_mean_ul', float),
    'circumference mean ll': ('circumference_mean_ll', float),
    'circumference sd max limits': ('circumference_sd_max', float),
    'cig. pdo': ('cig_pdo', float), 'cig. pdo ul': ('cig_pdo_ul', float),
    'cig. pdo ll': ('cig_pdo_ll', float),
    'tip ventilation (vf)': ('tip_ventilation', float),
    'tip ventilation (vf) ul': ('tip_ventilation_ul', float),
    'tip ventilation (vf) ll': ('tip_ventilation_ll', float),
    'tip ventilation (vf) sd max limit': ('tip_ventilation_sd_max', float),
    'rod length': ('tobacco_rod_length', float), 'cig length': ('cig_length', float),
    'ntm wt. mean': ('ntm_wt_mean', float), 'cig wt. sd max limit': ('cig_wt_sd_max', float),
    'filter pd': ('filter_pd', float), 'filter pd ul': ('filter_pd_ul', float),
    'filter pd ll': ('filter_pd_ll', float),
    'cig. hardness': ('cig_hardness', float), 'cig. hardness ul': ('cig_hardness_ul', float),
    'cig. hardness ll': ('cig_hardness_ll', float),
    'cig. corrected hardness': ('cig_corrected_hardness', float),
    'loose shorts mg/end max limits': ('loose_shorts_max', float),
    'no. of cut': ('c_plg', int), 'pluglength': ('plug_length', float),
    'filter weight': ('filter_weight', float),
    'c48 moisture': ('c48_moisture', float), 'c48 moisture ul': ('c48_moisture_ul', float),
    'c48 moisture ll': ('c48_moisture_ll', float),
    'maker moisture': ('maker_moisture', float), 'maker moisture ul': ('maker_moisture_ul', float),
    'maker moisture ll': ('maker_moisture_ll', float),
    'pack ov': ('pack_ov', float), 'pack ov ul': ('pack_ov_ul', float),
    'pack ov ll': ('pack_ov_ll', float),
    'ssi': ('ssi', float), 'ssi ul': ('ssi_ul', float), 'ssi ll': ('ssi_ll', float),
    'lamina cpi': ('lamina_cpi', float),
    'filling power': ('filling_power', float), 'filling power ul': ('filling_power_ul', float),
    'filling power ll': ('filling_power_ll', float),
    'pan % max limit': ('pan_pct_max', float),
    'filter desc': ('filter_desc', str), 'plug wrap cu': ('plug_wrap_cu', float),
    'tow': ('tow_used', str), 'target nic': ('target_nic', float),
}


def _tl_norm(h):
    """Normalize an Excel header: drop colons, collapse whitespace/newlines, lowercase."""
    return ' '.join(str(h).replace(':', ' ').split()).lower() if h is not None else ''


def _tl_code(v):
    """FG code cell → canonical 'NNNNNNNN.NN' string (Excel stores codes as floats)."""
    if v is None:
        return ''
    if isinstance(v, (int, float)):
        return '%.2f' % v
    return str(v).strip()


@bp.route('/targets-limits/upload', methods=['POST'])
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_EDIT, Permissions.MASTER_DATA_TARGETS_LIMITS)
def targets_limits_upload():
    """Upload a Targets & Limits Excel file and upsert fg_codes rows by FG code."""
    import openpyxl

    f = request.files.get('file')
    if not f or not f.filename or not f.filename.lower().endswith(('.xlsx', '.xlsm')):
        flash_error('Please choose an .xlsx or .xlsm file')
        return redirect(url_for('master_data.targets_limits'))

    try:
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    except Exception as e:
        flash_error(f'Could not read Excel file: {e}')
        return redirect(url_for('master_data.targets_limits'))

    sheet = next((s for s in ('Target & Limits for CGR8S', 'Targets & Limits') if s in wb.sheetnames),
                 wb.sheetnames[0])
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        flash_error('The sheet is empty')
        return redirect(url_for('master_data.targets_limits'))

    col = {_tl_norm(h): i for i, h in enumerate(headers) if h is not None}
    if 'fg code' not in col:
        flash_error(f'No "FG Code:" column found in sheet "{sheet}"')
        return redirect(url_for('master_data.targets_limits'))

    repo = FGCodeRepository(g.db)
    created = updated = skipped = 0
    errors = []

    for rownum, row in enumerate(rows, start=2):
        code = _tl_code(row[col['fg code']])
        if not code or code.lower() in ('none', 'nan'):
            skipped += 1
            continue

        data = {}
        for header, (field, typ) in _TL_FIELD_MAP.items():
            idx = col.get(header)
            if idx is None or idx >= len(row):
                continue
            v = row[idx]
            if v is None or (isinstance(v, str) and (not v.strip() or v.strip().startswith('#'))):
                continue  # blank or Excel error value → keep existing
            try:
                data[field] = typ(v) if typ is not str else str(v).strip()
            except (ValueError, TypeError):
                continue

        # ponytail: legacy rows imported via str(float) may store '.10' codes as '.1'
        fg = repo.get_by_code(code) or (code.endswith('0') and repo.get_by_code(code[:-1])) or None
        try:
            if fg:
                for field, value in data.items():
                    setattr(fg, field, value)
                updated += 1
            else:
                data['fg_code'] = code
                data['is_active'] = True
                repo.create(data)
                created += 1
        except Exception as e:
            errors.append(f'Row {rownum} ({code}): {e}')
            if len(errors) >= 10:
                break

    if errors:
        g.db.rollback()
        flash_error('Upload aborted: ' + '; '.join(errors[:3]) + (f' … and {len(errors)-3} more' if len(errors) > 3 else ''))
        return redirect(url_for('master_data.targets_limits'))

    g.db.commit()
    AuditLogger.log(AuditAction.UPDATE, 'FGCode', entity_id=None,
                    after_value={'upload': f.filename, 'sheet': sheet, 'created': created,
                                 'updated': updated, 'skipped': skipped},
                    module='master_data')
    flash_success(f'Targets & Limits uploaded: {created} created, {updated} updated, {skipped} skipped')
    return redirect(url_for('master_data.targets_limits'))


@bp.route('/targets-limits')
@require_auth
@require_any_permissions(Permissions.MASTER_DATA_VIEW, Permissions.MASTER_DATA_TARGETS_LIMITS)
def targets_limits():
    page, per_page = paginate_args(request.args)
    repo = FGCodeRepository(g.db)
    result = repo.get_paginated(page=page, per_page=per_page,
                                search=request.args.get('q', ''),
                                search_fields=['fg_code', 'brand', 'blend', 'family_name'])
    return render_template('master_data/targets_limits.html', **result)
