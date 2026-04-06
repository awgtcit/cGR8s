"""Verify Excel freeze panes are set correctly (columns up to SKU frozen)."""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from playwright.sync_api import sync_playwright
import openpyxl

BASE = "http://localhost:5053"

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()

    page.goto(f"{BASE}/qa/data-grid", wait_until="networkidle", timeout=15000)
    if '/qa/data-grid' not in page.url:
        page.goto(f"{BASE}/login/test-bypass", wait_until="networkidle", timeout=15000)
        page.wait_for_timeout(1000)
        page.goto(f"{BASE}/qa/data-grid", wait_until="networkidle", timeout=15000)
        page.wait_for_timeout(1000)

    # Download Excel
    with page.expect_download() as dl_info:
        page.locator('a:has-text("Excel")').click()
    download = dl_info.value
    path = f"test-results/{download.suggested_filename}"
    download.save_as(path)
    print(f"Downloaded: {path}")

    browser.close()

# Check freeze panes
wb = openpyxl.load_workbook(path)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n{name}:")
    print(f"  freeze_panes = {ws.freeze_panes}")
    # Find SKU column
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if 'SKU' in headers:
        sku_idx = headers.index('SKU')
        print(f"  SKU at column {sku_idx + 1} ({openpyxl.utils.get_column_letter(sku_idx + 1)})")
        expected_freeze = f"{openpyxl.utils.get_column_letter(sku_idx + 2)}2"
        assert str(ws.freeze_panes) == expected_freeze, \
            f"Expected freeze at {expected_freeze}, got {ws.freeze_panes}"
        print(f"  ✅ Frozen up to SKU (freeze at {expected_freeze})")
    else:
        print(f"  No SKU column found, headers: {headers[:10]}...")

print("\n✅ All sheets verified!")
