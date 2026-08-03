"""Import cGr8s.xlsm production history into the grid tables so the NPL Data Grid
populates. Maps the Production Data + QA Analysis sheets into:
  process_orders, process_order_key_variables, target_weight_results,
  npl_inputs, npl_results, qa_analysis.

Rows whose SKU is absent from fg_codes are skipped (NOT-NULL FK). Duplicate
(PO, date) pairs get a "/n" suffix so no matched row is dropped. Idempotent
guard: aborts if process_orders already has rows.

  set CGR8S_SRC_XLSM to a readable copy if the root file is OneDrive-locked.
  .venv/Scripts/python.exe scripts/import_production_history.py
"""
import os
import sys
import uuid
import shutil
import tempfile
import urllib.parse
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env"),
            override=True)

import openpyxl
from sqlalchemy import create_engine, text

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.getenv("CGR8S_SRC_XLSM") or os.path.join(PROJECT_ROOT, "cGr8s.xlsm")


def engine():
    drv = os.getenv("DB_DRIVER", "ODBC Driver 17 for SQL Server")
    return create_engine(
        f"mssql+pyodbc://{urllib.parse.quote_plus(os.getenv('DB_USER'))}:"
        f"{urllib.parse.quote_plus(os.getenv('DB_PASSWORD'))}@{os.getenv('DB_SERVER')},"
        f"{os.getenv('DB_PORT', '1433')}/{os.getenv('DB_NAME')}"
        f"?driver={urllib.parse.quote_plus(drv)}", connect_args={"timeout": 60})


def f(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def s(v, n=None):
    if v is None:
        return None
    t = str(v).strip()
    return (t[:n] if n else t) or None


def load_wb():
    try:
        tmp = os.path.join(tempfile.gettempdir(), "cGr8s_hist.xlsm")
        shutil.copyfile(SRC, tmp)
    except PermissionError:
        tmp = SRC
    return openpyxl.load_workbook(tmp, read_only=True, data_only=True)


def batch_insert(conn, table, cols, rows, chunk=1000):
    if not rows:
        return
    collist = ", ".join(cols)
    vallist = ", ".join(f":{c}" for c in cols)
    sql = text(f"INSERT INTO {table} ({collist}) VALUES ({vallist})")
    clean = [{k: row[k] for k in cols} for row in rows]  # project to exactly cols
    for i in range(0, len(clean), chunk):
        conn.execute(sql, clean[i:i + chunk])


def main():
    eng = engine()
    with eng.connect() as c:
        n = c.execute(text("SELECT COUNT(*) FROM process_orders")).scalar()
        if n:
            print(f"ABORT: process_orders already has {n} rows. "
                  "Purge first or this would duplicate.")
            return
        fg_map = {r[0]: r[1] for r in c.execute(text("SELECT fg_code, id FROM fg_codes"))}
    print(f"fg_codes loaded: {len(fg_map)}")

    wb = load_wb()
    pd = wb["Production Data"]
    qa = wb["QA Analysis"]

    # QA rows keyed by RID (col A)
    qa_by_rid = {}
    for r in qa.iter_rows(min_row=3, values_only=True):
        if r[0] is None:
            continue
        qa_by_rid[r[0]] = r

    po_rows, kv_rows, tw_rows, ni_rows, nr_rows, qa_rows = [], [], [], [], [], []
    seen_pairs = defaultdict(int)
    skipped_sku = skipped_date = 0

    for r in pd.iter_rows(min_row=3, values_only=True):
        if r[0] is None:
            continue
        sku = s(r[5], 50)
        fg_id = fg_map.get(sku)
        if not fg_id:
            skipped_sku += 1
            continue
        pdate = r[2]
        if not hasattr(pdate, "year"):
            skipped_date += 1
            continue

        po_num = s(r[4], 45) or "N/A"
        key = (po_num, str(pdate)[:10])
        seen_pairs[key] += 1
        if seen_pairs[key] > 1:
            po_num = f"{po_num}/{seen_pairs[key]}"

        po_id = str(uuid.uuid4())
        ni_id = str(uuid.uuid4())
        tw_id = str(uuid.uuid4())
        nr_id = str(uuid.uuid4())

        po_rows.append({"id": po_id, "fg": fg_id, "num": po_num[:50], "pdate": pdate,
                        "status": s(r[87], 30) or "Done"})
        kv_rows.append({"id": str(uuid.uuid4()), "po": po_id,
                        "n_bld": f(r[13]), "p_cu": f(r[14]), "t_vnt": f(r[15]),
                        "f_pd": f(r[16]), "m_ip": f(r[17]), "alpha": f(r[18]),
                        "beta": f(r[19]), "gamma": f(r[20]), "delta": f(r[21]),
                        "n_tgt": f(r[23]), "cd": pdate})
        tw_rows.append({"id": tw_id, "po": po_id, "cd": pdate,
                        "s1": f(r[42]), "s2": f(r[43]), "td": f(r[44]), "fp": f(r[45]),
                        "ps1": f(r[46]), "ps2": f(r[47]), "pt": f(r[48]),
                        "tnd": f(r[49]), "tfp": f(r[50]),
                        "wdry": f(r[51]), "wtob": f(r[52]), "wcig": f(r[53]),
                        "wntm": f(r[24]), "tw": f(r[53]),
                        "in_nbld": f(r[13]), "in_pcu": f(r[14]), "in_tvnt": f(r[15]),
                        "in_fpd": f(r[16]), "in_mip": f(r[17]), "in_a": f(r[18]),
                        "in_b": f(r[19]), "in_g": f(r[20]), "in_d": f(r[21]),
                        "in_nt": f(r[23])})
        ni_rows.append({"id": ni_id, "po": po_id, "cd": pdate,
                        "t_iss": f(r[26]), "t_un": f(r[27]), "l_dst": f(r[28]),
                        "l_win": f(r[29]), "l_flr": f(r[30]), "l_srt": f(r[31]),
                        "l_dt": f(r[32]), "r_mkg": f(r[33]), "r_pkg": f(r[34]),
                        "r_ndt": f(r[35]), "n_w": f(r[36]), "n_mc": f(r[37]),
                        "n_cg": f(r[38]), "t_usd": f(r[39]), "m_dsp": f(r[40]),
                        "m_dst": f(r[41])})
        nr_rows.append({"id": nr_id, "po": po_id, "ni": ni_id, "cd": pdate,
                        "tac": f(r[54]), "ttc": f(r[55]), "pct": f(r[56]), "kg": f(r[57])})

        q = qa_by_rid.get(r[0])
        if q is not None:
            qa_rows.append({"id": str(uuid.uuid4()), "po": po_id, "nr": nr_id, "tw": tw_id,
                            "cd": pdate, "status": s(q[41], 30) or "Done",
                            "pack_ov": f(q[13]), "lamina_cpi": f(q[14]),
                            "filling_power": f(q[15]), "filling_power_corr": f(q[16]),
                            "maker_moisture": f(q[17]), "ssi": f(q[18]), "pan_pct": f(q[19]),
                            "total_cig_length": f(q[20]), "circumference_mean": f(q[21]),
                            "circumference_sd": f(q[22]), "cig_dia": f(q[23]),
                            "tobacco_weight_mean": f(q[24]), "tobacco_weight_sd": f(q[25]),
                            "tip_vf": f(q[26]), "tip_vf_sd": f(q[27]),
                            "filter_pd_mean": f(q[28]), "filter_weight": f(q[29]),
                            "plug_wrap_cu": f(q[30]), "tow": s(q[31], 100),
                            "cig_wt_mean": f(q[32]), "cig_wt_sd": f(q[33]),
                            "cig_pdo": f(q[34]), "cig_hardness": f(q[35]),
                            "cig_corr_hardness": f(q[36]), "loose_shorts": f(q[37]),
                            "plug_length": f(q[38]), "mc": s(q[39], 50), "company": s(q[40], 100)})
    wb.close()

    print(f"prepared: PO {len(po_rows)}, KV {len(kv_rows)}, TW {len(tw_rows)}, "
          f"NPLin {len(ni_rows)}, NPLres {len(nr_rows)}, QA {len(qa_rows)}")
    print(f"skipped (SKU not in fg_codes): {skipped_sku}, (no prod date): {skipped_date}, "
          f"PO+date suffixed: {sum(v - 1 for v in seen_pairs.values() if v > 1)}")

    with eng.begin() as conn:
        batch_insert(conn, "process_orders",
                     ["id", "fg_code_id", "process_order_number", "process_date", "status",
                      "created_at", "updated_at", "is_deleted", "row_version"],
                     [{**p, "created_at": p["pdate"], "updated_at": p["pdate"],
                       "is_deleted": 0, "row_version": 1,
                       "fg_code_id": p["fg"], "process_order_number": p["num"],
                       "process_date": p["pdate"]} for p in po_rows])
        batch_insert(conn, "process_order_key_variables",
                     ["id", "process_order_id", "n_bld", "p_cu", "t_vnt", "f_pd", "m_ip",
                      "alpha", "beta", "gamma", "delta", "n_tgt",
                      "created_at", "updated_at", "row_version"],
                     [{**k, "process_order_id": k["po"], "created_at": k["cd"],
                       "updated_at": k["cd"], "row_version": 1} for k in kv_rows])
        batch_insert(conn, "target_weight_results",
                     ["id", "process_order_id", "calculated_at", "created_at", "updated_at",
                      "stage1_dilution", "stage2_dilution", "total_dilution", "filtration_pct",
                      "stage1_pacifying_nicotine_demand", "stage2_pacifying_nicotine_demand",
                      "total_pacifying_nicotine_demand", "total_filtration_pct",
                      "total_nicotine_demand", "tw", "w_dry", "w_tob", "w_cig", "w_ntm",
                      "input_n_bld", "input_p_cu", "input_t_vnt", "input_f_pd", "input_m_ip",
                      "input_alpha", "input_beta", "input_gamma", "input_delta", "input_n_tgt"],
                     [{"id": t["id"], "process_order_id": t["po"], "calculated_at": t["cd"],
                       "created_at": t["cd"], "updated_at": t["cd"],
                       "stage1_dilution": t["s1"], "stage2_dilution": t["s2"],
                       "total_dilution": t["td"], "filtration_pct": t["fp"],
                       "stage1_pacifying_nicotine_demand": t["ps1"],
                       "stage2_pacifying_nicotine_demand": t["ps2"],
                       "total_pacifying_nicotine_demand": t["pt"],
                       "total_filtration_pct": t["tfp"], "total_nicotine_demand": t["tnd"],
                       "tw": t["tw"], "w_dry": t["wdry"], "w_tob": t["wtob"],
                       "w_cig": t["wcig"], "w_ntm": t["wntm"],
                       "input_n_bld": t["in_nbld"], "input_p_cu": t["in_pcu"],
                       "input_t_vnt": t["in_tvnt"], "input_f_pd": t["in_fpd"],
                       "input_m_ip": t["in_mip"], "input_alpha": t["in_a"],
                       "input_beta": t["in_b"], "input_gamma": t["in_g"],
                       "input_delta": t["in_d"], "input_n_tgt": t["in_nt"]} for t in tw_rows])
        batch_insert(conn, "npl_inputs",
                     ["id", "process_order_id", "created_at", "updated_at",
                      "t_iss", "t_un", "l_dst", "l_win", "l_flr", "l_srt", "l_dt",
                      "r_mkg", "r_pkg", "r_ndt", "n_w", "n_mc", "n_cg", "t_usd", "m_dsp", "m_dst"],
                     [{**i, "process_order_id": i["po"],
                       "created_at": i["cd"], "updated_at": i["cd"]} for i in ni_rows])
        batch_insert(conn, "npl_results",
                     ["id", "process_order_id", "npl_input_id", "calculated_at",
                      "created_at", "updated_at", "verified", "tac", "ttc", "npl_pct", "npl_kg"],
                     [{"id": n["id"], "process_order_id": n["po"], "npl_input_id": n["ni"],
                       "calculated_at": n["cd"], "created_at": n["cd"], "updated_at": n["cd"],
                       "verified": 1, "tac": n["tac"], "ttc": n["ttc"],
                       "npl_pct": n["pct"], "npl_kg": n["kg"]} for n in nr_rows])
        batch_insert(conn, "qa_analysis",
                     ["id", "process_order_id", "npl_result_id", "target_weight_result_id",
                      "status", "created_at", "updated_at", "analyzed_at", "row_version",
                      "pack_ov", "lamina_cpi", "filling_power", "filling_power_corr",
                      "maker_moisture", "ssi", "pan_pct", "total_cig_length",
                      "circumference_mean", "circumference_sd", "cig_dia",
                      "tobacco_weight_mean", "tobacco_weight_sd", "tip_vf", "tip_vf_sd",
                      "filter_pd_mean", "filter_weight", "plug_wrap_cu", "tow",
                      "cig_wt_mean", "cig_wt_sd", "cig_pdo", "cig_hardness",
                      "cig_corr_hardness", "loose_shorts", "plug_length", "mc", "company"],
                     [{**q, "process_order_id": q["po"], "npl_result_id": q["nr"],
                       "target_weight_result_id": q["tw"], "created_at": q["cd"],
                       "updated_at": q["cd"], "analyzed_at": q["cd"],
                       "row_version": 1} for q in qa_rows])

    with eng.connect() as conn:
        for t in ("process_orders", "process_order_key_variables", "target_weight_results",
                  "npl_inputs", "npl_results", "qa_analysis"):
            print(f"  {t:30} {conn.execute(text(f'SELECT COUNT(*) FROM {t}')).scalar()}")


if __name__ == "__main__":
    main()
