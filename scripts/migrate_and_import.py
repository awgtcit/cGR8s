"""
Schema migration + full Excel data import.
1. ALTER fg_codes to add target/limit columns
2. CREATE tables: machines, skus, tobacco_blend_analysis
3. Import all data from Data.xlsx
"""
import os
import sys
import io
import msoffcrypto
import openpyxl
from sqlalchemy import create_engine, text

# ── Connection ────────────────────────────────────────────────────────────
ENGINE_URL = "mssql+pyodbc://sa:Admin%40123@172.50.35.75/cGR8s?driver=ODBC+Driver+17+for+SQL+Server"
engine = create_engine(ENGINE_URL)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'cigr8s vb', 'ALW', 'Data', 'Data.xlsx')
PASSWORD = 'ALW'


def decrypt_workbook():
    """Decrypt the password-protected Excel file and return openpyxl workbook."""
    with open(EXCEL_PATH, 'rb') as f:
        msfile = msoffcrypto.OfficeFile(f)
        msfile.load_key(password=PASSWORD)
        buf = io.BytesIO()
        msfile.decrypt(buf)
        buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)


# ── Step 1: Add target/limit columns to fg_codes ──────────────────────────
NEW_FG_COLUMNS = [
    ('family_name', 'NVARCHAR(100)'),
    ('circumference_mean', 'FLOAT'), ('circumference_mean_ul', 'FLOAT'),
    ('circumference_mean_ll', 'FLOAT'), ('circumference_sd_max', 'FLOAT'),
    ('cig_pdo', 'FLOAT'), ('cig_pdo_ul', 'FLOAT'), ('cig_pdo_ll', 'FLOAT'),
    ('tip_ventilation', 'FLOAT'), ('tip_ventilation_ul', 'FLOAT'),
    ('tip_ventilation_ll', 'FLOAT'), ('tip_ventilation_sd_max', 'FLOAT'),
    ('ntm_wt_mean', 'FLOAT'), ('cig_wt_sd_max', 'FLOAT'),
    ('filter_pd', 'FLOAT'), ('filter_pd_ul', 'FLOAT'), ('filter_pd_ll', 'FLOAT'),
    ('cig_hardness', 'FLOAT'), ('cig_hardness_ul', 'FLOAT'), ('cig_hardness_ll', 'FLOAT'),
    ('cig_corrected_hardness', 'FLOAT'), ('loose_shorts_max', 'FLOAT'),
    ('filter_weight', 'FLOAT'),
    ('c48_moisture', 'FLOAT'), ('c48_moisture_ul', 'FLOAT'), ('c48_moisture_ll', 'FLOAT'),
    ('maker_moisture', 'FLOAT'), ('maker_moisture_ul', 'FLOAT'), ('maker_moisture_ll', 'FLOAT'),
    ('pack_ov', 'FLOAT'), ('pack_ov_ul', 'FLOAT'), ('pack_ov_ll', 'FLOAT'),
    ('ssi', 'FLOAT'), ('ssi_ul', 'FLOAT'), ('ssi_ll', 'FLOAT'),
    ('lamina_cpi', 'FLOAT'),
    ('filling_power', 'FLOAT'), ('filling_power_ul', 'FLOAT'),
    ('filling_power_ll', 'FLOAT'), ('filling_power_corrected_ul', 'FLOAT'),
    ('pan_pct_max', 'FLOAT'),
    ('filter_desc', 'NVARCHAR(200)'), ('plug_wrap_cu', 'FLOAT'), ('target_nic', 'FLOAT'),
]


def alter_fg_codes(conn):
    """Add missing columns to fg_codes table."""
    existing = {row[0] for row in conn.execute(text(
        "SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'fg_codes'"
    ))}
    added = 0
    for col_name, col_type in NEW_FG_COLUMNS:
        if col_name not in existing:
            conn.execute(text(f"ALTER TABLE fg_codes ADD {col_name} {col_type} NULL"))
            added += 1
    print(f"  fg_codes: added {added} new columns (skipped {len(NEW_FG_COLUMNS) - added} existing)")


# ── Step 2: Create new tables ─────────────────────────────────────────────
def create_new_tables(conn):
    """Create machines, skus, tobacco_blend_analysis tables if not exists."""
    existing_tables = {row[0] for row in conn.execute(text(
        "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_TYPE = 'BASE TABLE'"
    ))}

    if 'machines' not in existing_tables:
        conn.execute(text("""
            CREATE TABLE machines (
                id NVARCHAR(36) PRIMARY KEY,
                machine_code NVARCHAR(50) NOT NULL UNIQUE,
                description NVARCHAR(200) NULL,
                plant NVARCHAR(50) NULL,
                format_type NVARCHAR(100) NULL,
                is_active BIT DEFAULT 1 NOT NULL,
                created_at DATETIME2 DEFAULT GETDATE(),
                updated_at DATETIME2 DEFAULT GETDATE(),
                created_by NVARCHAR(100) NULL,
                updated_by NVARCHAR(100) NULL,
                is_deleted BIT DEFAULT 0 NOT NULL,
                deleted_at DATETIME2 NULL,
                deleted_by NVARCHAR(100) NULL,
                row_version INT DEFAULT 1 NOT NULL
            )
        """))
        print("  Created table: machines")
    else:
        print("  Table machines already exists")

    if 'skus' not in existing_tables:
        conn.execute(text("""
            CREATE TABLE skus (
                id NVARCHAR(36) PRIMARY KEY,
                sku_code NVARCHAR(50) NOT NULL UNIQUE,
                description NVARCHAR(300) NULL,
                nicotine FLOAT NULL,
                ventilation FLOAT NULL,
                pd_code NVARCHAR(50) NULL,
                cig_code NVARCHAR(50) NULL,
                is_active BIT DEFAULT 1 NOT NULL,
                created_at DATETIME2 DEFAULT GETDATE(),
                updated_at DATETIME2 DEFAULT GETDATE(),
                created_by NVARCHAR(100) NULL,
                updated_by NVARCHAR(100) NULL,
                is_deleted BIT DEFAULT 0 NOT NULL,
                deleted_at DATETIME2 NULL,
                deleted_by NVARCHAR(100) NULL,
                row_version INT DEFAULT 1 NOT NULL
            )
        """))
        conn.execute(text("CREATE INDEX ix_skus_cig_code ON skus (cig_code)"))
        print("  Created table: skus")
    else:
        print("  Table skus already exists")

    if 'tobacco_blend_analysis' not in existing_tables:
        conn.execute(text("""
            CREATE TABLE tobacco_blend_analysis (
                id NVARCHAR(36) PRIMARY KEY,
                period_year INT NULL,
                period_month INT NULL,
                blend_name NVARCHAR(100) NOT NULL,
                nic_wet FLOAT NULL,
                nic_dry FLOAT NULL,
                dispatch_moisture FLOAT NULL,
                is_active BIT DEFAULT 1 NOT NULL,
                created_at DATETIME2 DEFAULT GETDATE(),
                updated_at DATETIME2 DEFAULT GETDATE(),
                created_by NVARCHAR(100) NULL,
                updated_by NVARCHAR(100) NULL,
                is_deleted BIT DEFAULT 0 NOT NULL,
                deleted_at DATETIME2 NULL,
                deleted_by NVARCHAR(100) NULL,
                row_version INT DEFAULT 1 NOT NULL
            )
        """))
        conn.execute(text("CREATE INDEX ix_tba_blend_name ON tobacco_blend_analysis (blend_name)"))
        print("  Created table: tobacco_blend_analysis")
    else:
        print("  Table tobacco_blend_analysis already exists")


# ── Step 3: Import Excel data ─────────────────────────────────────────────

def safe_float(val):
    """Convert value to float, return None if not numeric."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def safe_str(val, max_len=None):
    """Convert value to trimmed string, return None if empty."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s


def import_targets_limits(conn, wb):
    """Import/update FG codes with target & limit data from 'Targets & Limits' sheet."""
    ws = wb['Targets & Limits']
    rows = list(ws.iter_rows(min_row=3, values_only=True))  # Skip header rows
    print(f"  Targets & Limits: {len(rows)} data rows")

    updated = 0
    not_found = 0
    for row in rows:
        fg_code_val = safe_str(row[1])
        if not fg_code_val:
            continue

        # Build update dict – map Excel columns to DB columns
        updates = {
            'cig_code': safe_str(row[2], 50),
            'blend_code': safe_str(row[3], 50),
            'filter_code': safe_str(row[4], 50),
            'blend': safe_str(row[5], 100),
            'brand': safe_str(row[6], 100),
            'format': safe_str(row[7], 50),
            'family_name': safe_str(row[8], 100),
            'fg_gtin': safe_str(row[9], 50),
            'blend_gtin': safe_str(row[10], 50),
            'circumference_mean': safe_float(row[11]),
            'circumference_mean_ul': safe_float(row[12]),
            'circumference_mean_ll': safe_float(row[13]),
            'circumference_sd_max': safe_float(row[14]),
            'cig_pdo': safe_float(row[15]),
            'cig_pdo_ul': safe_float(row[16]),
            'cig_pdo_ll': safe_float(row[17]),
            'tip_ventilation': safe_float(row[18]),
            'tip_ventilation_ul': safe_float(row[19]),
            'tip_ventilation_ll': safe_float(row[20]),
            'tip_ventilation_sd_max': safe_float(row[21]),
            'tobacco_rod_length': safe_float(row[22]),
            'cig_length': safe_float(row[23]),
            'ntm_wt_mean': safe_float(row[24]),
            'cig_wt_sd_max': safe_float(row[25]),
            'filter_pd': safe_float(row[26]),
            'filter_pd_ul': safe_float(row[27]),
            'filter_pd_ll': safe_float(row[28]),
            'cig_hardness': safe_float(row[29]),
            'cig_hardness_ul': safe_float(row[30]),
            'cig_hardness_ll': safe_float(row[31]),
            'cig_corrected_hardness': safe_float(row[32]),
            'loose_shorts_max': safe_float(row[33]),
            'filter_length': safe_float(row[34]),
            'c_plg': safe_float(row[35]),
            'plug_length': safe_float(row[36]),
            'filter_weight': safe_float(row[37]),
            'c48_moisture': safe_float(row[38]),
            'c48_moisture_ul': safe_float(row[39]),
            'c48_moisture_ll': safe_float(row[40]),
            'maker_moisture': safe_float(row[41]),
            'maker_moisture_ul': safe_float(row[42]),
            'maker_moisture_ll': safe_float(row[43]),
            'pack_ov': safe_float(row[44]),
            'pack_ov_ul': safe_float(row[45]),
            'pack_ov_ll': safe_float(row[46]),
            'ssi': safe_float(row[47]),
            'ssi_ul': safe_float(row[48]),
            'ssi_ll': safe_float(row[49]),
            'lamina_cpi': safe_float(row[50]),
            'filling_power': safe_float(row[51]),
            'filling_power_ul': safe_float(row[52]),
            'filling_power_ll': safe_float(row[53]),
            'filling_power_corrected_ul': safe_float(row[54]),
            'pan_pct_max': safe_float(row[55]),
            'filter_desc': safe_str(row[56], 200),
            'plug_wrap_cu': safe_float(row[57]),
            'tow_used': safe_str(row[58], 100),
            'target_nic': safe_float(row[59]),
        }

        # Remove None values to avoid overwriting existing data with NULL
        updates = {k: v for k, v in updates.items() if v is not None}

        if not updates:
            continue

        # Build parameterized SET clause
        set_parts = [f"{k} = :v_{k}" for k in updates]
        params = {f"v_{k}": v for k, v in updates.items()}
        params['fg_code'] = fg_code_val

        result = conn.execute(
            text(f"UPDATE fg_codes SET {', '.join(set_parts)} WHERE fg_code = :fg_code AND is_deleted = 0"),
            params
        )
        if result.rowcount > 0:
            updated += 1
        else:
            not_found += 1

    print(f"    Updated: {updated}, Not found in DB: {not_found}")


def import_blends(conn, wb):
    """Import missing blends from Data sheet (cols 18-21)."""
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    imported = 0
    skipped = 0
    for row in rows:
        blend_code = safe_str(row[17], 50)  # col 18 (0-indexed: 17)
        if not blend_code or blend_code == 'Blend Code':
            continue

        description = safe_str(row[18], 500)
        blend_gtin = safe_str(row[19], 50)
        plant = safe_str(row[20], 50)

        # Check if exists
        exists = conn.execute(
            text("SELECT 1 FROM blend_master WHERE blend_code = :bc"),
            {'bc': blend_code}
        ).fetchone()

        if exists:
            skipped += 1
            continue

        import uuid
        conn.execute(
            text("""INSERT INTO blend_master (id, blend_code, blend_name, blend_gtin, description, is_active, row_version, created_at, updated_at, is_deleted)
                     VALUES (:id, :bc, :bn, :bg, :desc, 1, 1, GETDATE(), GETDATE(), 0)"""),
            {
                'id': str(uuid.uuid4()),
                'bc': blend_code,
                'bn': description or blend_code,
                'bg': blend_gtin,
                'desc': f"Plant: {plant}" if plant else None,
            }
        )
        imported += 1

    print(f"  Blends: imported {imported}, skipped {skipped} existing")


def import_machines(conn, wb):
    """Import machines from Data sheet (cols 23-26)."""
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    imported = 0
    for row in rows:
        machine_code = safe_str(row[22], 50)  # col 23 (0-indexed: 22)
        if not machine_code or machine_code == 'Machine No':
            continue

        description = safe_str(row[23], 200)
        plant = safe_str(row[24], 50)
        format_type = safe_str(row[25], 100)

        # Check if exists
        exists = conn.execute(
            text("SELECT 1 FROM machines WHERE machine_code = :mc"),
            {'mc': machine_code}
        ).fetchone()

        if exists:
            continue

        import uuid
        conn.execute(
            text("""INSERT INTO machines (id, machine_code, description, plant, format_type, is_active, row_version)
                     VALUES (:id, :mc, :desc, :plant, :fmt, 1, 1)"""),
            {
                'id': str(uuid.uuid4()),
                'mc': machine_code,
                'desc': description,
                'plant': plant,
                'fmt': format_type,
            }
        )
        imported += 1

    print(f"  Machines: imported {imported}")


def import_skus(conn, wb):
    """Import SKUs from Data sheet (cols 28-33)."""
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    imported = 0
    batch = []
    for row in rows:
        sku_code = safe_str(row[27], 50)  # col 28 (0-indexed: 27)
        if not sku_code or sku_code == 'SKU':
            continue

        import uuid
        batch.append({
            'id': str(uuid.uuid4()),
            'sku': sku_code,
            'desc': safe_str(row[28], 300),
            'nic': safe_float(row[29]),
            'vent': safe_float(row[30]),
            'pd': safe_str(row[31], 50),
            'cig': safe_str(row[32], 50),
        })

    # Bulk insert, skip duplicates
    for rec in batch:
        exists = conn.execute(
            text("SELECT 1 FROM skus WHERE sku_code = :sc"),
            {'sc': rec['sku']}
        ).fetchone()
        if exists:
            continue

        conn.execute(
            text("""INSERT INTO skus (id, sku_code, description, nicotine, ventilation, pd_code, cig_code, is_active, row_version)
                     VALUES (:id, :sku, :desc, :nic, :vent, :pd, :cig, 1, 1)"""),
            rec
        )
        imported += 1

    print(f"  SKUs: imported {imported}")


def import_tobacco_blend_analysis(conn, wb):
    """Import tobacco blend analysis from Data sheet (cols 4-8)."""
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    imported = 0
    for row in rows:
        period_raw = row[3]  # col 4 (0-indexed: 3) - "YY MM" format
        blend_name = safe_str(row[4], 100)
        if not blend_name or blend_name == 'Blend':
            continue

        period_year = None
        period_month = None
        if period_raw is not None:
            ps = str(period_raw).strip()
            parts = ps.split()
            if len(parts) >= 2:
                try:
                    period_year = int(parts[0]) + 2000 if int(parts[0]) < 100 else int(parts[0])
                    period_month = int(parts[1])
                except (ValueError, TypeError):
                    pass

        nic_wet = safe_float(row[5])
        nic_dry = safe_float(row[6])
        dispatch_moisture = safe_float(row[7])

        import uuid
        conn.execute(
            text("""INSERT INTO tobacco_blend_analysis
                     (id, period_year, period_month, blend_name, nic_wet, nic_dry, dispatch_moisture, is_active, row_version)
                     VALUES (:id, :py, :pm, :bn, :nw, :nd, :dm, 1, 1)"""),
            {
                'id': str(uuid.uuid4()),
                'py': period_year,
                'pm': period_month,
                'bn': blend_name,
                'nw': nic_wet,
                'nd': nic_dry,
                'dm': dispatch_moisture,
            }
        )
        imported += 1

    print(f"  Tobacco Blend Analysis: imported {imported}")


def import_lookups(conn, wb):
    """Import Size/CU, KP Tolerances, Plug Length/Cuts as lookups."""
    ws = wb['Data']
    rows = list(ws.iter_rows(min_row=3, values_only=True))
    import uuid

    imported = 0

    # Size/CU (cols 1-2, 0-indexed: 0-1)
    for row in rows:
        size_code = safe_str(row[0], 50)
        cu_val = safe_str(row[1], 50)
        if not size_code or size_code == 'Size':
            continue

        exists = conn.execute(
            text("SELECT 1 FROM lookups WHERE category = 'size_cu' AND code = :code"),
            {'code': size_code}
        ).fetchone()
        if exists:
            continue

        conn.execute(
            text("""INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at)
                     VALUES (:id, 'size_cu', :code, :dn, :so, 1, GETDATE(), GETDATE())"""),
            {'id': str(uuid.uuid4()), 'code': size_code, 'dn': f"{size_code} = {cu_val}", 'so': imported + 1}
        )
        imported += 1

    # KP Tolerance (cols 10-13, 0-indexed: 9-12)
    kp_count = 0
    for row in rows:
        kp_param = safe_str(row[9], 100)
        if not kp_param or kp_param == 'Parameter':
            continue

        exists = conn.execute(
            text("SELECT 1 FROM lookups WHERE category = 'kp_tolerance' AND code = :code"),
            {'code': kp_param}
        ).fetchone()
        if exists:
            continue

        stage1 = safe_str(row[10], 50) or ''
        stage2 = safe_str(row[11], 50) or ''
        stage3 = safe_str(row[12], 50) or ''
        display = f"S1: {stage1}, S2: {stage2}, S3: {stage3}"

        conn.execute(
            text("""INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at)
                     VALUES (:id, 'kp_tolerance', :code, :dn, :so, 1, GETDATE(), GETDATE())"""),
            {'id': str(uuid.uuid4()), 'code': kp_param, 'dn': display, 'so': kp_count + 1}
        )
        kp_count += 1
        imported += 1

    # Plug Length/Cuts (cols 15-16, 0-indexed: 14-15)
    pc_count = 0
    for row in rows:
        plug_len = safe_str(row[14], 50)
        cuts = safe_str(row[15], 50)
        if not plug_len or plug_len == 'Plug Length':
            continue

        exists = conn.execute(
            text("SELECT 1 FROM lookups WHERE category = 'plug_length_cuts' AND code = :code"),
            {'code': plug_len}
        ).fetchone()
        if exists:
            continue

        conn.execute(
            text("""INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at)
                     VALUES (:id, 'plug_length_cuts', :code, :dn, :so, 1, GETDATE(), GETDATE())"""),
            {'id': str(uuid.uuid4()), 'code': plug_len, 'dn': f"Plug {plug_len}mm = {cuts} cuts", 'so': pc_count + 1}
        )
        pc_count += 1
        imported += 1

    print(f"  Lookups (Size/CU, KP Tolerance, Plug/Cuts): imported {imported}")


# ── Main ──────────────────────────────────────────────────────────────────
def main():
    print("=" * 60)
    print("cGR8s Schema Migration & Excel Data Import")
    print("=" * 60)

    print("\n[1/5] Adding columns to fg_codes...")
    with engine.begin() as conn:
        alter_fg_codes(conn)

    print("\n[2/5] Creating new tables...")
    with engine.begin() as conn:
        create_new_tables(conn)

    print("\n[3/5] Decrypting Excel workbook...")
    wb = decrypt_workbook()
    print(f"  Sheets: {wb.sheetnames}")

    print("\n[4/5] Importing data from 'Targets & Limits' sheet...")
    with engine.begin() as conn:
        import_targets_limits(conn, wb)

    print("\n[5/5] Importing data from 'Data' sheet...")
    with engine.begin() as conn:
        import_blends(conn, wb)
    with engine.begin() as conn:
        import_machines(conn, wb)
    with engine.begin() as conn:
        import_skus(conn, wb)
    with engine.begin() as conn:
        import_tobacco_blend_analysis(conn, wb)
    with engine.begin() as conn:
        import_lookups(conn, wb)

    print("\n" + "=" * 60)
    print("Import complete! Verifying counts...")
    with engine.connect() as conn:
        for table in ['fg_codes', 'blend_master', 'machines', 'skus', 'tobacco_blend_analysis', 'lookups']:
            count = conn.execute(text(f"SELECT COUNT(*) FROM {table}")).scalar()
            print(f"  {table}: {count} rows")
    print("=" * 60)


if __name__ == '__main__':
    main()
