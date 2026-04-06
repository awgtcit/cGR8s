"""
Import data from VB application Excel files into cGR8s MSSQL database.

Sources:
  - cigr8s vb/cGr8s/Report Excel/Data/cGr8s.xlsm  (GTIN sheet → fg_codes, blend_master)
  - cigr8s vb/program files/ALW/Constants.xlsx       (calibration constants)
  - cigr8s vb/cGr8s/Report Excel/Data/cGr8s.xlsm  (Production Data → process_orders)
"""
import os
import sys
import uuid
from datetime import datetime, timezone

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from dotenv import load_dotenv
load_dotenv(override=True)

import openpyxl
from app import create_app
from app.database import get_engine, get_session, Base
from app.models.fg_code import FGCode
from app.models.blend_master import BlendMaster
from app.models.calibration_constant import CalibrationConstant
from app.models.physical_parameter import PhysicalParameter
from app.models.lookup import Lookup
from app.models.base import generate_uuid

# Paths
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
VB_DIR = os.path.join(PROJECT_ROOT, 'cigr8s vb')
CGRS_XLSM = os.path.join(VB_DIR, 'cGr8s', 'Report Excel', 'Data', 'cGr8s.xlsm')
CONSTANTS_XLSX = os.path.join(VB_DIR, 'program files', 'ALW', 'Constants.xlsx')


def import_gtin_master(session):
    """Import GTIN sheet → fg_codes table + blend_master table."""
    print("\n=== Importing GTIN Master Data ===")
    wb = openpyxl.load_workbook(CGRS_XLSM, read_only=True, data_only=True)
    ws = wb['GTIN']

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    # Map columns by header name
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    imported_fg = 0
    skipped_fg = 0
    blends_seen = set()
    imported_blends = 0

    for row in data_rows:
        if not row or not row[col_map.get('Material', 1)]:
            continue

        material = str(row[col_map.get('Material', 1)]).strip()
        description = str(row[col_map.get('Material Description', 2)] or '').strip()
        pack_barcode = row[col_map.get('Pack Barcode', 3)]
        outer_barcode = row[col_map.get('Outer Barcode', 4)]
        carton_barcode = row[col_map.get('Carton Barcode', 5)]

        # Skip dust/waste/non-product entries
        if not material or material == 'None':
            continue

        # Determine format from description
        fmt = _extract_format(description)
        brand = _extract_brand(description)

        # Check if fg_code already exists
        existing = session.query(FGCode).filter(FGCode.fg_code == material).first()
        if existing:
            skipped_fg += 1
            continue

        fg = FGCode(
            id=generate_uuid(),
            fg_code=material,
            brand=brand,
            fg_gtin=str(pack_barcode) if pack_barcode else None,
            format=fmt,
            is_active=True,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        )
        session.add(fg)
        imported_fg += 1

    session.flush()
    print(f"  FG Codes: {imported_fg} imported, {skipped_fg} skipped (existing)")

    wb.close()
    return imported_fg


def import_production_data_skus(session):
    """
    Import unique SKUs from Production Data sheet into fg_codes + calibration + blends.
    This fills in details like blend_code, blend_gtin, cig_code, format, etc.
    """
    print("\n=== Importing SKU Details from Production Data ===")
    wb = openpyxl.load_workbook(CGRS_XLSM, read_only=True, data_only=True)
    ws = wb['Production Data']

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    headers = rows[0]
    data_rows = rows[1:]

    # Build header map from the actual header row (row index 1 in the data)
    # Headers are: RID, QR Date, Prod Date, NPL Date, Process Order, SKU, SKU Desc, ...
    header_names = [
        'RID', 'QR_Date', 'Prod_Date', 'NPL_Date', 'Process_Order',
        'SKU', 'SKU_Desc', 'SKU_GTN', 'Blend_Code', 'Blend_Desc',
        'Blend_GTIN', 'Cig_Code', 'UID', 'NBLD', 'PCU', 'VF', 'FPD',
        'MIP', 'Alpha', 'Beta', 'Gamma', 'Delta', 'NC', 'NT', 'NTM',
        'TW', 'TISS', 'TUN', 'RW1', 'RW2', 'RW3', 'RW4', 'RW5',
        'WM', 'WP', 'WQ', 'NW', 'NMC', 'NCG', 'T_USD', 'M_DSP', 'M_DST',
        'S1', 'S2', 'TD', 'F', 'NS1', 'NS2', 'NTD', 'NF', 'NTDRY',
        'W_DRY', 'W_TOB', 'W_CIG', 'TAC', 'TTC', 'NPL_PCT', 'NPL_KG',
        'FG_Blend', 'Pack_OV', 'Lamina_CPI', 'Filling_Power',
        'Filling_Power_Corr', 'Maker_Moisture', 'SSI', 'PAN_PCT',
        'Total_Cig_Length', 'Circumference_Mean', 'Circumference_SD',
        'Cig_Dia', 'Tobacco_Weight_Mean', 'Tobacco_Weight_SD',
        'TIP_VF', 'TIP_VF_SD', 'Filter_PD_Mean', 'Filter_Weight',
        'Plug_Wrap_CU', 'TOW', 'Cig_Wt_Mean', 'Cig_Wt_SD',
        'Cig_PDO', 'Cig_Hardness', 'Cig_Corr_Hardness', 'Loose_Shorts',
        'Plug_Length', 'MC', 'Company', 'Status', 'Next_Row'
    ]

    skus_seen = {}
    blends_seen = {}
    updated_fg = 0
    new_blends = 0
    new_calibrations = 0

    for row in data_rows:
        if not row or not row[0]:
            continue

        # Extract values by position
        def val(idx):
            return row[idx] if idx < len(row) else None

        sku = str(val(5) or '').strip()
        if not sku or sku == 'None':
            continue

        # Only process first occurrence of each SKU
        if sku in skus_seen:
            continue
        skus_seen[sku] = True

        sku_desc = str(val(6) or '').strip()
        sku_gtn = val(7)
        blend_code = str(val(8) or '').strip()
        blend_desc = str(val(9) or '').strip()
        blend_gtin = val(10)
        cig_code = str(val(11) or '').strip()
        nbld = val(13)
        pcu = val(14)
        vf = val(15)
        fpd = val(16)
        mip = val(17)
        alpha = val(18)
        beta = val(19)
        gamma = val(20)
        delta = val(21)
        plug_length = val(84)
        tow = val(77)
        total_cig_length = val(66)

        # Extract format from description
        fmt = _extract_format(sku_desc)
        brand = _extract_brand(sku_desc)

        # Update or create FG Code
        fg = session.query(FGCode).filter(FGCode.fg_code == sku).first()
        if fg:
            # Update with richer data from production
            fg.brand = brand or fg.brand
            fg.fg_gtin = str(sku_gtn) if sku_gtn else fg.fg_gtin
            fg.format = fmt or fg.format
            fg.blend_code = blend_code or fg.blend_code
            fg.blend = blend_desc or fg.blend
            fg.blend_gtin = str(blend_gtin) if blend_gtin else fg.blend_gtin
            fg.cig_code = cig_code or fg.cig_code
            fg.plug_length = _safe_float(plug_length) or fg.plug_length
            fg.tow_used = str(tow) if tow else fg.tow_used
            fg.cig_length = _safe_float(total_cig_length) or fg.cig_length
            fg.updated_at = datetime.now(timezone.utc)
            updated_fg += 1
        else:
            fg = FGCode(
                id=generate_uuid(),
                fg_code=sku,
                brand=brand,
                fg_gtin=str(sku_gtn) if sku_gtn else None,
                format=fmt,
                blend_code=blend_code,
                blend=blend_desc,
                blend_gtin=str(blend_gtin) if blend_gtin else None,
                cig_code=cig_code,
                plug_length=_safe_float(plug_length),
                tow_used=str(tow) if tow else None,
                cig_length=_safe_float(total_cig_length),
                is_active=True,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(fg)
            session.flush()
            updated_fg += 1

        # Create/update blend master
        if blend_code and blend_code not in blends_seen:
            blends_seen[blend_code] = True
            existing_blend = session.query(BlendMaster).filter(
                BlendMaster.blend_code == blend_code
            ).first()
            if not existing_blend:
                blend = BlendMaster(
                    id=generate_uuid(),
                    blend_code=blend_code,
                    blend_name=blend_desc,
                    blend_gtin=str(blend_gtin) if blend_gtin else None,
                    n_bld=_safe_float(nbld),
                    is_active=True,
                    created_at=datetime.now(timezone.utc),
                    updated_at=datetime.now(timezone.utc),
                )
                session.add(blend)
                new_blends += 1

        # Create calibration constants for this FG code
        existing_cal = session.query(CalibrationConstant).filter(
            CalibrationConstant.fg_code_id == fg.id
        ).first()
        if not existing_cal:
            cal = CalibrationConstant(
                id=generate_uuid(),
                fg_code_id=fg.id,
                alpha=_safe_float(alpha),
                beta=_safe_float(beta),
                gamma=_safe_float(gamma),
                delta=_safe_float(delta),
                n_tgt=_safe_float(val(23)),  # NT column
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(cal)
            new_calibrations += 1

        # Create physical parameters
        existing_phys = session.query(PhysicalParameter).filter(
            PhysicalParameter.fg_code_id == fg.id
        ).first()
        if not existing_phys:
            phys = PhysicalParameter(
                id=generate_uuid(),
                fg_code_id=fg.id,
                p_cu=_safe_float(pcu),
                t_vnt=_safe_float(vf),
                f_pd=_safe_float(fpd),
                m_ip=_safe_float(mip),
                cig_length=_safe_float(total_cig_length),
                plug_length=_safe_float(plug_length),
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(phys)

    session.flush()
    print(f"  FG Codes updated/created: {updated_fg}")
    print(f"  Blend Masters created: {new_blends}")
    print(f"  Calibration Constants created: {new_calibrations}")

    wb.close()
    return updated_fg


def import_global_constants(session):
    """Import global constants from Constants.xlsx → system_config / lookups."""
    print("\n=== Importing Global Constants ===")
    wb = openpyxl.load_workbook(CONSTANTS_XLSX, read_only=True, data_only=True)
    ws = wb['Data']

    # Global constants are in columns G-H (indices 6-7)
    # Row structure: Name in col G, Value in col H
    global_constants = {
        'Alpha': 10,
        'Beta': -0.043,
        'Delta': -0.056,
        'Dust Moisture': 9,
        'Max VF': 65,
        'Min PD': 220,
        'Max PD': 450,
        'NCG': 10000,
    }

    # Read actual values from the file
    rows = list(ws.iter_rows(min_row=1, max_row=20, min_col=7, max_col=8, values_only=True))
    for row in rows:
        if row[0] and row[1] is not None:
            name = str(row[0]).strip()
            value = row[1]
            global_constants[name] = value

    imported = 0
    for name, value in global_constants.items():
        existing = session.query(Lookup).filter(
            Lookup.category == 'global_constant',
            Lookup.code == name
        ).first()
        if not existing:
            lookup = Lookup(
                id=generate_uuid(),
                category='global_constant',
                code=name,
                display_name=f"{name}: {value}",
                sort_order=imported,
                is_active=True,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(lookup)
            imported += 1

    # Import format-specific constants as lookups
    # Read format constants from columns A-E
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    formats_imported = 0
    for row in rows:
        if not row[0]:
            continue
        fmt = str(row[0]).strip()
        plug_len = row[1]
        condition = row[2]
        criteria = str(row[3]).strip() if row[3] else ''
        constant = row[4]

        if not fmt or fmt == 'None':
            continue

        # Skip formula rows (plug_len contains a formula string, not a number)
        if plug_len is not None and not isinstance(plug_len, (int, float)):
            continue

        lookup_code = f"{fmt}|{plug_len}|{criteria}"
        # Truncate to fit column limit (50 chars)
        if len(lookup_code) > 50:
            lookup_code = lookup_code[:50]
        existing = session.query(Lookup).filter(
            Lookup.category == 'filtration_constant',
            Lookup.code == lookup_code
        ).first()
        if not existing:
            display = f"{fmt} PL={plug_len} {criteria}={constant}"
            if len(display) > 100:
                display = display[:100]
            lookup = Lookup(
                id=generate_uuid(),
                category='filtration_constant',
                code=lookup_code,
                display_name=display,
                sort_order=formats_imported,
                is_active=True,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(lookup)
            formats_imported += 1

    session.flush()
    print(f"  Global constants: {imported} imported")
    print(f"  Filtration constants: {formats_imported} imported")

    wb.close()
    return imported + formats_imported


def import_format_lookups(session):
    """Import cigarette format types as lookup data."""
    print("\n=== Importing Format Lookups ===")
    formats = [
        'KS10SE', 'KS20FP', 'KS20RC', 'KS20SE', 'KS21', 'KS25', 'KS27',
        'NS/SS', 'NS20DS', 'NS20SE', 'QS', 'QS20RC', 'SK20SP', 'SKS/SL',
        'SL20SE', 'SS20SE', 'KS (NIC<0.2)'
    ]

    imported = 0
    for i, fmt in enumerate(formats):
        existing = session.query(Lookup).filter(
            Lookup.category == 'cigarette_format',
            Lookup.code == fmt
        ).first()
        if not existing:
            lookup = Lookup(
                id=generate_uuid(),
                category='cigarette_format',
                code=fmt,
                display_name=fmt,
                sort_order=i,
                is_active=True,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            )
            session.add(lookup)
            imported += 1

    session.flush()
    print(f"  Cigarette formats: {imported} imported")
    return imported


def _extract_format(description):
    """Extract cigarette format from SKU description."""
    if not description:
        return None
    desc = description.upper().strip()
    # Match known format suffixes
    format_patterns = [
        'NS20SE', 'NS20DS', 'KS20SE', 'KS20FP', 'KS20RC', 'KS10SE',
        'QS20RC', 'SK20SP', 'SL20SE', 'SS20SE', 'KS21', 'KS25', 'KS27',
        'NS/SS', 'SKS/SL', 'QS',
    ]
    for fmt in format_patterns:
        if fmt in desc:
            return fmt
    # Fallback - try to extract last token
    parts = desc.split()
    if parts:
        last = parts[-1]
        if any(c.isdigit() for c in last):
            return last
    return None


def _extract_brand(description):
    """Extract brand name from SKU description."""
    if not description:
        return None
    desc = description.strip()
    # Remove format suffix
    format_patterns = [
        'NS20SE', 'NS20DS', 'KS20SE', 'KS20FP', 'KS20RC', 'KS10SE',
        'QS20RC', 'SK20SP', 'SL20SE', 'SS20SE', 'KS21', 'KS25', 'KS27',
        'NS/SS', 'SKS/SL', 'QS',
    ]
    brand = desc.upper()
    for fmt in format_patterns:
        brand = brand.replace(fmt, '').strip()
    # Clean up extra spaces
    brand = ' '.join(brand.split())
    # Title case
    if brand:
        return brand.title()
    return desc.strip()


def _safe_float(val):
    """Safely convert to float, returning None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def main():
    """Run all imports."""
    print("=" * 60)
    print("cGR8s VB Data Import")
    print("=" * 60)

    # Create Flask app for database context
    app = create_app(os.getenv('FLASK_ENV', 'development'))

    with app.app_context():
        engine = get_engine()
        # Disable SQL echo for performance
        engine.echo = False
        # Ensure tables exist
        Base.metadata.create_all(engine)

        session = get_session()
        try:
            # 1. Import GTIN master data (fg_codes base)
            import_gtin_master(session)
            session.commit()
            print("  [Stage 1 committed]")

            # 2. Import SKU details from Production Data (enriches fg_codes + creates blends/calibration)
            import_production_data_skus(session)
            session.commit()
            print("  [Stage 2 committed]")

            # 3. Import global constants
            import_global_constants(session)
            session.commit()
            print("  [Stage 3 committed]")

            # 4. Import format lookups
            import_format_lookups(session)
            session.commit()
            print("  [Stage 4 committed]")
            print("\n" + "=" * 60)
            print("Import completed successfully!")
            print("=" * 60)

            # Print summary
            fg_count = session.query(FGCode).filter(FGCode.is_deleted == False).count()
            blend_count = session.query(BlendMaster).filter(BlendMaster.is_deleted == False).count()
            cal_count = session.query(CalibrationConstant).count()
            phys_count = session.query(PhysicalParameter).count()
            lookup_count = session.query(Lookup).count()

            print(f"\nDatabase Summary:")
            print(f"  FG Codes:              {fg_count}")
            print(f"  Blend Masters:         {blend_count}")
            print(f"  Calibration Constants: {cal_count}")
            print(f"  Physical Parameters:   {phys_count}")
            print(f"  Lookups:               {lookup_count}")

        except Exception as e:
            session.rollback()
            print(f"\nERROR: {e}")
            import traceback
            traceback.print_exc()
        finally:
            session.close()


if __name__ == '__main__':
    main()
