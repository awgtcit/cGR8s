"""FULL RESET: purge master + transactional data, reload master from Data.xlsx.

Keeps engine constants intact (formula_constants, gamma_constants) and seeds a
global tobacco_constant into system_config. Run AFTER scripts/backup_db.py.

  .venv/Scripts/python.exe scripts/reset_and_load_from_data.py

Source: <project root>/Data.xlsx  (password ALW), sheets 'Targets & Limits' + 'Data'.
"""
import os
import sys
import io
import uuid
import shutil
import tempfile
import urllib.parse

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from dotenv import load_dotenv
load_dotenv(override=True)

import msoffcrypto
import openpyxl
from sqlalchemy import create_engine, text

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# Env override lets us point at a pre-copied file when the OneDrive-synced
# original is locked against Python's open() (Excel/OneDrive sharing mode).
DATA_XLSX = os.getenv("CGR8S_DATA_XLSX") or os.path.join(PROJECT_ROOT, "Data.xlsx")
PASSWORD = "ALW"
TOBACCO_CONSTANT = "0.99620799"

# ── purge order: children first, parents last. Constants NOT purged. ────────
PURGE_ORDER = [
    # transactional
    "qa_updates", "batch_job_items", "optimizer_results", "optimizer_inputs",
    "qa_analysis", "npl_results", "npl_inputs", "optimizer_runs",
    "target_weight_results", "process_order_key_variables", "reports",
    "batch_jobs", "audit_logs", "master_data_change_log", "process_orders",
    # fg_codes FK children
    "optimizer_limits", "calibration_constants", "physical_parameters",
    "product_versions",
    # master
    "tobacco_blend_analysis", "skus", "machines", "blend_master", "lookups",
    "fg_codes",
]


def engine():
    drv = os.getenv("DB_DRIVER", "ODBC Driver 17 for SQL Server")
    return create_engine(
        f"mssql+pyodbc://{urllib.parse.quote_plus(os.getenv('DB_USER'))}:"
        f"{urllib.parse.quote_plus(os.getenv('DB_PASSWORD'))}@{os.getenv('DB_SERVER')},"
        f"{os.getenv('DB_PORT', '1433')}/{os.getenv('DB_NAME')}"
        f"?driver={urllib.parse.quote_plus(drv)}",
        connect_args={"timeout": 30},
    )


def safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_str(v, n=None):
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    return s[:n] if n else s


def load_workbook():
    """Copy (avoids OneDrive lock), decrypt with password, return workbook."""
    tmp_enc = os.path.join(tempfile.gettempdir(), "cGR8s_Data_enc.xlsx")
    try:
        shutil.copyfile(DATA_XLSX, tmp_enc)
    except PermissionError:
        # source is locked; assume DATA_XLSX already points at a readable copy
        tmp_enc = DATA_XLSX
    with open(tmp_enc, "rb") as f:
        off = msoffcrypto.OfficeFile(f)
        off.load_key(password=PASSWORD)
        buf = io.BytesIO()
        off.decrypt(buf)
    buf.seek(0)
    return openpyxl.load_workbook(buf, read_only=True, data_only=True)


# ── purge ───────────────────────────────────────────────────────────────
def purge(conn):
    print("\n[1/4] Purging (constants kept)...")
    total = 0
    for t in PURGE_ORDER:
        n = conn.execute(text(f"DELETE FROM [{t}]")).rowcount
        total += max(n, 0)
        print(f"  - {t:30} {n if n and n > 0 else 0:>6}")
    print(f"  purged ~{total} rows")


# ── fg_codes INSERT from Targets & Limits ─────────────────────────────────
FG_COLS = [
    ("cig_code", 2, "s", 50), ("blend_code", 3, "s", 50), ("filter_code", 4, "s", 50),
    ("blend", 5, "s", 100), ("brand", 6, "s", 100), ("format", 7, "s", 50),
    ("family_name", 8, "s", 100), ("fg_gtin", 9, "s", 50), ("blend_gtin", 10, "s", 50),
    ("circumference_mean", 11, "f", 0), ("circumference_mean_ul", 12, "f", 0),
    ("circumference_mean_ll", 13, "f", 0), ("circumference_sd_max", 14, "f", 0),
    ("cig_pdo", 15, "f", 0), ("cig_pdo_ul", 16, "f", 0), ("cig_pdo_ll", 17, "f", 0),
    ("tip_ventilation", 18, "f", 0), ("tip_ventilation_ul", 19, "f", 0),
    ("tip_ventilation_ll", 20, "f", 0), ("tip_ventilation_sd_max", 21, "f", 0),
    ("tobacco_rod_length", 22, "f", 0), ("cig_length", 23, "f", 0),
    ("ntm_wt_mean", 24, "f", 0), ("cig_wt_sd_max", 25, "f", 0),
    ("filter_pd", 26, "f", 0), ("filter_pd_ul", 27, "f", 0), ("filter_pd_ll", 28, "f", 0),
    ("cig_hardness", 29, "f", 0), ("cig_hardness_ul", 30, "f", 0), ("cig_hardness_ll", 31, "f", 0),
    ("cig_corrected_hardness", 32, "f", 0), ("loose_shorts_max", 33, "f", 0),
    ("filter_length", 34, "f", 0), ("c_plg", 35, "i", 0), ("plug_length", 36, "f", 0),
    ("filter_weight", 37, "f", 0), ("c48_moisture", 38, "f", 0),
    ("c48_moisture_ul", 39, "f", 0), ("c48_moisture_ll", 40, "f", 0),
    ("maker_moisture", 41, "f", 0), ("maker_moisture_ul", 42, "f", 0),
    ("maker_moisture_ll", 43, "f", 0), ("pack_ov", 44, "f", 0),
    ("pack_ov_ul", 45, "f", 0), ("pack_ov_ll", 46, "f", 0),
    ("ssi", 47, "f", 0), ("ssi_ul", 48, "f", 0), ("ssi_ll", 49, "f", 0),
    ("lamina_cpi", 50, "f", 0), ("filling_power", 51, "f", 0),
    ("filling_power_ul", 52, "f", 0), ("filling_power_ll", 53, "f", 0),
    ("filling_power_corrected_ul", 54, "f", 0), ("pan_pct_max", 55, "f", 0),
    ("filter_desc", 56, "s", 200), ("plug_wrap_cu", 57, "f", 0),
    ("tow_used", 58, "s", 100), ("target_nic", 59, "f", 0),
]


def load_fg_codes(conn, wb):
    ws = wb["Targets & Limits"]
    inserted = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        fg_code = safe_str(row[1], 50)
        if not fg_code or fg_code == "FG Code":
            continue
        rec = {"id": str(uuid.uuid4()), "fg_code": fg_code}
        for name, idx, typ, ln in FG_COLS:
            if idx >= len(row):
                rec[name] = None
                continue
            v = row[idx]
            if typ == "s":
                rec[name] = safe_str(v, ln)
            elif typ == "i":
                fv = safe_float(v)
                rec[name] = int(fv) if fv is not None else None
            else:
                rec[name] = safe_float(v)
        cols = ["id", "fg_code"] + [c[0] for c in FG_COLS]
        collist = ", ".join(cols) + ", is_active, created_at, updated_at, is_deleted, row_version"
        vallist = ", ".join(f":{c}" for c in cols) + ", 1, GETDATE(), GETDATE(), 0, 1"
        try:
            conn.execute(text(f"INSERT INTO fg_codes ({collist}) VALUES ({vallist})"), rec)
            inserted += 1
        except Exception as e:
            skipped += 1
            if skipped <= 3:
                print(f"    ! skip {fg_code}: {str(e)[:90]}")
    print(f"  fg_codes: inserted {inserted}, skipped {skipped}")


# ── Data-sheet blocks (verified column indices) ───────────────────────────
def load_data_blocks(conn, wb):
    ws = wb["Data"]
    rows = list(ws.iter_rows(min_row=3, values_only=True))

    # blend_master  (R/S/T = 17/18/19)
    n = 0
    for row in rows:
        bc = safe_str(row[17], 50)
        if not bc or bc == "Blend Code":
            continue
        if conn.execute(text("SELECT 1 FROM blend_master WHERE blend_code=:c"), {"c": bc}).fetchone():
            continue
        desc = safe_str(row[18], 100)
        conn.execute(text(
            "INSERT INTO blend_master (id, blend_code, blend_name, blend_gtin, description, is_active, row_version, created_at, updated_at, is_deleted) "
            "VALUES (:id,:bc,:bn,:bg,:d,1,1,GETDATE(),GETDATE(),0)"),
            {"id": str(uuid.uuid4()), "bc": bc, "bn": desc or bc,
             "bg": safe_str(row[19], 50), "d": None})
        n += 1
    print(f"  blend_master: {n}")

    # machines  (W/X/Y/Z = 22/23/24/25)
    n = 0
    for row in rows:
        mc = safe_str(row[22], 50)
        if not mc or mc == "Machine No":
            continue
        if conn.execute(text("SELECT 1 FROM machines WHERE machine_code=:c"), {"c": mc}).fetchone():
            continue
        conn.execute(text(
            "INSERT INTO machines (id, machine_code, description, plant, format_type, is_active, row_version) "
            "VALUES (:id,:mc,:d,:p,:f,1,1)"),
            {"id": str(uuid.uuid4()), "mc": mc, "d": safe_str(row[23], 200),
             "p": safe_str(row[24], 50), "f": safe_str(row[25], 100)})
        n += 1
    print(f"  machines: {n}")

    # skus  (AB..AG = 27/28/29/30/31/32)
    n = 0
    for row in rows:
        sc = safe_str(row[27], 50)
        if not sc or sc == "SKU":
            continue
        if conn.execute(text("SELECT 1 FROM skus WHERE sku_code=:c"), {"c": sc}).fetchone():
            continue
        conn.execute(text(
            "INSERT INTO skus (id, sku_code, description, nicotine, ventilation, pd_code, cig_code, is_active, row_version) "
            "VALUES (:id,:sku,:d,:nic,:v,:pd,:cig,1,1)"),
            {"id": str(uuid.uuid4()), "sku": sc, "d": safe_str(row[28], 300),
             "nic": safe_float(row[29]), "v": safe_float(row[30]),
             "pd": safe_str(row[31], 50), "cig": safe_str(row[32], 50)})
        n += 1
    print(f"  skus: {n}")

    # tobacco_blend_analysis  (D/E/F/G/H = 3/4/5/6/7)
    n = 0
    for row in rows:
        bn = safe_str(row[4], 100)
        if not bn or bn == "Blend":
            continue
        py = pm = None
        if row[3] is not None:
            parts = str(row[3]).strip().split()
            if len(parts) >= 2:
                try:
                    py = int(parts[0]) + 2000 if int(parts[0]) < 100 else int(parts[0])
                    pm = int(parts[1])
                except (ValueError, TypeError):
                    pass
        conn.execute(text(
            "INSERT INTO tobacco_blend_analysis (id, period_year, period_month, blend_name, nic_wet, nic_dry, dispatch_moisture, is_active, row_version) "
            "VALUES (:id,:py,:pm,:bn,:nw,:nd,:dm,1,1)"),
            {"id": str(uuid.uuid4()), "py": py, "pm": pm, "bn": bn,
             "nw": safe_float(row[5]), "nd": safe_float(row[6]), "dm": safe_float(row[7])})
        n += 1
    print(f"  tobacco_blend_analysis: {n}")

    # lookups: size_cu (A/B=0/1), kp_tolerance (J..M=9..12), plug_length_cuts (O/P=14/15)
    n = 0
    for row in rows:
        code = safe_str(row[0], 50)
        if not code or code == "Size":
            continue
        if conn.execute(text("SELECT 1 FROM lookups WHERE category='size_cu' AND code=:c"), {"c": code}).fetchone():
            continue
        conn.execute(text(
            "INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at) "
            "VALUES (:id,'size_cu',:c,:dn,:so,1,GETDATE(),GETDATE())"),
            {"id": str(uuid.uuid4()), "c": code, "dn": f"{code} = {safe_str(row[1], 50)}", "so": n + 1})
        n += 1
    for row in rows:
        kp = safe_str(row[9], 100)
        if not kp or kp == "Parameter":
            continue
        if conn.execute(text("SELECT 1 FROM lookups WHERE category='kp_tolerance' AND code=:c"), {"c": kp}).fetchone():
            continue
        disp = f"S1: {safe_str(row[10],50) or ''}, S2: {safe_str(row[11],50) or ''}, S3: {safe_str(row[12],50) or ''}"
        conn.execute(text(
            "INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at) "
            "VALUES (:id,'kp_tolerance',:c,:dn,:so,1,GETDATE(),GETDATE())"),
            {"id": str(uuid.uuid4()), "c": kp, "dn": disp, "so": n + 1})
        n += 1
    for row in rows:
        pl = safe_str(row[14], 50)
        if not pl or pl == "Plug Length":
            continue
        if conn.execute(text("SELECT 1 FROM lookups WHERE category='plug_length_cuts' AND code=:c"), {"c": pl}).fetchone():
            continue
        conn.execute(text(
            "INSERT INTO lookups (id, category, code, display_name, sort_order, is_active, created_at, updated_at) "
            "VALUES (:id,'plug_length_cuts',:c,:dn,:so,1,GETDATE(),GETDATE())"),
            {"id": str(uuid.uuid4()), "c": pl, "dn": f"Plug {pl}mm = {safe_str(row[15],50)} cuts", "so": n + 1})
        n += 1
    print(f"  lookups (size_cu + kp_tolerance + plug_length_cuts): {n}")


def seed_tobacco_constant(conn):
    exists = conn.execute(text("SELECT 1 FROM system_config WHERE config_key='tobacco_constant'")).fetchone()
    if exists:
        conn.execute(text("UPDATE system_config SET config_value=:v, updated_at=GETDATE() WHERE config_key='tobacco_constant'"),
                     {"v": TOBACCO_CONSTANT})
        print(f"  system_config.tobacco_constant updated = {TOBACCO_CONSTANT}")
    else:
        conn.execute(text(
            "INSERT INTO system_config (id, config_key, config_value, description, is_sensitive, created_at, updated_at) "
            "VALUES (:id,'tobacco_constant',:v,:d,0,GETDATE(),GETDATE())"),
            {"id": str(uuid.uuid4()), "v": TOBACCO_CONSTANT,
             "d": "Global tobacco density constant; divides W_TOB (cell-trace global-last fallback)"})
        print(f"  system_config.tobacco_constant inserted = {TOBACCO_CONSTANT}")


def main():
    print("=" * 60)
    print("cGR8s FULL RESET + reload from Data.xlsx")
    print("=" * 60)
    wb = load_workbook()
    print("workbook decrypted; sheets:", wb.sheetnames)
    eng = engine()
    with eng.begin() as conn:
        purge(conn)
        print("\n[2/4] Loading fg_codes (Targets & Limits)...")
        load_fg_codes(conn, wb)
        print("\n[3/4] Loading Data-sheet blocks...")
        load_data_blocks(conn, wb)
        print("\n[4/4] Seeding tobacco constant...")
        seed_tobacco_constant(conn)
    print("\nDONE. Final counts:")
    with eng.connect() as conn:
        for t in ("fg_codes", "skus", "blend_master", "machines",
                  "tobacco_blend_analysis", "lookups", "process_orders",
                  "formula_constants", "gamma_constants"):
            print(f"  {t:26} {conn.execute(text(f'SELECT COUNT(*) FROM [{t}]')).scalar()}")


if __name__ == "__main__":
    main()
